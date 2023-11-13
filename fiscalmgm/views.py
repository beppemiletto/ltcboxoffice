from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.utils import timezone
from django.utils.formats import date_format
from datetime import datetime, timedelta
import pytz
from store.models import Event
from tickets.models import Ticket
from .models import Ingresso, Report, EventFiscalData
from .forms import FiscalDataForm

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.drawing.image import Image

import os
import re

# Create your views here.
def user_not_allowed(request):
    messages.error(request,"Sei entrato con utente non di staff. Non puoi operare in Amministrazione (Cassa e SIAE).")
    return render(request, 'accounts/login.html')

@login_required(login_url='login')
def fiscalmgm_main(request):
    if request.user.is_staff:
        now = datetime.now(pytz.timezone('Europe/Rome')) + timedelta(hours=6)
        # time_diff = timedelta(days= 365)
        past_events = Event.objects.filter(date_time__lte=now).filter(show__is_active=True).order_by('-date_time')
        
        paginator_past = Paginator(past_events, 10)
        page = request.GET.get('page')
        paged_events = paginator_past.get_page(page)
        
        context = {
            'past_events' : paged_events,
        }
        return render(request, 'fiscalmgm/event_list.html', context)
    
@login_required(login_url='login')
def siae(request, event_id):
    if request.user.is_staff:
        curr_event =  Event.objects.get(id = event_id)
        now = datetime.now(pytz.timezone('Europe/Rome')) + timedelta(hours=6)
        past_events = Event.objects.filter(date_time__lte=now).filter(show__is_active=True).order_by('-date_time')
        del now
        # print("{} - number {} will be the one for SIAE reports".format(curr_event,event_id))
        save_path = 'media/siae_reports'
        evnt_date_code = curr_event.date_time.strftime('%Y%m%d')
        evnt_show_slug = curr_event.show.slug
        evnt_show_code = curr_event.show.shw_code


        # Retrieve if exist the record of fiscl data - UNIQUE for Event
        # Otherwise create it taking most updated values as default.
        try:
            fiscal_data = EventFiscalData.objects.get(event_id = curr_event.pk)
        except:
            fiscal_data_old = EventFiscalData.objects.last()
            fiscal_data = EventFiscalData()
            fiscal_data.event = curr_event


            # calcola l'incasso da prevendita
            full_price:float = curr_event.price_full
            redu_price:float = curr_event.price_reduced
            prices = [0.0 , redu_price, full_price] 
            all_tickets = Ticket.objects.filter(event_id = curr_event.pk)
            presold_tickets = all_tickets.filter(sell_mode = "W")
            presold_income: float = 0.0 
            for presold_ticket in presold_tickets:
                presold_income += prices[presold_ticket.price]

            fiscal_data.collection_presell = presold_income
            if fiscal_data_old is not None:
                fiscal_data.cash_begin_200euro = fiscal_data_old.cash_end_200euro
                fiscal_data.cash_begin_100euro = fiscal_data_old.cash_end_100euro
                fiscal_data.cash_begin_50euro = fiscal_data_old.cash_end_50euro
                fiscal_data.cash_begin_20euro = fiscal_data_old.cash_end_20euro
                fiscal_data.cash_begin_10euro = fiscal_data_old.cash_end_10euro
                fiscal_data.cash_begin_5euro = fiscal_data_old.cash_end_5euro
                fiscal_data.cash_begin_2euro = fiscal_data_old.cash_end_2euro
                fiscal_data.cash_begin_1euro = fiscal_data_old.cash_end_1euro
                fiscal_data.cash_begin_50cent = fiscal_data_old.cash_end_50cent
                fiscal_data.cash_begin_20cent = fiscal_data_old.cash_end_20cent
                fiscal_data.cash_begin_10cent = fiscal_data_old.cash_end_10cent
                if fiscal_data_old.intero_ticket_serie_2 != 'nd':
                    fiscal_data.intero_ticket_serie_1 = fiscal_data_old.intero_ticket_serie_2
                    fiscal_data.intero_ticket_start_number_1 = fiscal_data_old.intero_ticket_end_number_2 + 1
                else:
                    fiscal_data.intero_ticket_serie_1 = fiscal_data_old.intero_ticket_serie_1
                    fiscal_data.intero_ticket_start_number_1 = fiscal_data_old.intero_ticket_end_number_1 + 1
                if fiscal_data_old.ridotto_ticket_serie_2 != 'nd':
                    fiscal_data.ridotto_ticket_serie_1 = fiscal_data_old.ridotto_ticket_serie_2
                    fiscal_data.ridotto_ticket_start_number_1 = fiscal_data_old.ridotto_ticket_end_number_2 + 1
                else:
                    fiscal_data.ridotto_ticket_serie_1 = fiscal_data_old.ridotto_ticket_serie_1
                    fiscal_data.ridotto_ticket_start_number_1 = fiscal_data_old.ridotto_ticket_end_number_1 + 1
                if fiscal_data_old.gratuito_ticket_serie_2 != 'nd':
                    fiscal_data.gratuito_ticket_serie_1 = fiscal_data_old.gratuito_ticket_serie_2
                    fiscal_data.gratuito_ticket_start_number_1 = fiscal_data_old.gratuito_ticket_end_number_2 + 1
                else:
                    fiscal_data.gratuito_ticket_serie_1 = fiscal_data_old.gratuito_ticket_serie_1
                    fiscal_data.gratuito_ticket_start_number_1 = fiscal_data_old.gratuito_ticket_end_number_1 + 1
            fiscal_data.save()
        report_done: bool = fiscal_data.printed


        if request.method == 'POST':

            

            form = FiscalDataForm(request.POST, instance= fiscal_data)
            if form.is_valid():
                form.save()
            else:
                context = {
                'event' : curr_event,
                'form' : form,

            }
                return render(request, 'fiscalmgm/event_reportdata.html', context)

                


            font = {
            'normal' : Font(name='Arial', size=10,bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000'),
            'normal07' : Font(name='Arial', size=7,bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000'),
            'normal09' : Font(name='Arial', size=9,bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000'),
            'normal11' : Font(name='Arial', size=11,bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000'),
            'bold' : Font(name='Arial', size=10,bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000'),
            'bold09' : Font(name='Arial', size=9,bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000'),
            'bold11' : Font(name='Arial', size=11,bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000'),
            'bold13' : Font(name='Arial', size=13,bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000'),
            }
            alignment = {            
                'l_b' : Alignment(horizontal='left', vertical='bottom', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0),
                'l_c' : Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0),
                'c_c' : Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0),
                'r_c' : Alignment(horizontal='right', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0),
                }
            
            localtz = timezone.get_current_timezone()
            date_time = curr_event.date_time.astimezone(localtz)        
            esercizio_sociale_inizio = datetime.fromisoformat(f'{curr_event.date_time.year}-01-01')
            esercizio_sociale_fine = datetime.fromisoformat(f'{curr_event.date_time.year}-12-31')
        
            # Prepare data ingressi
            ingressi = {}
            ingressi['globale'] = Ingresso.objects.filter(event = curr_event)
            ingressi['interi'] = ingressi['globale'].filter(price=curr_event.price_full)
            ingressi['ridotti'] = ingressi['globale'].filter(price=curr_event.price_reduced)
            ingressi['omaggio'] = ingressi['globale'].filter(price=0.0)
            # print(ingressi['globale'].count(), ingressi['interi'].count(),ingressi['ridotti'].count(), ingressi['omaggio'].count())
            introiti = {}
            introiti['interi'] = ingressi['interi'].count() * curr_event.price_full
            introiti['ridotti'] = ingressi['ridotti'].count() * curr_event.price_reduced
            introiti['globale'] = introiti['interi'] + introiti['ridotti']

            # prepare data for SIAE titoli report
                    
            titoli = []
            if fiscal_data.intero_ticket_start_number_1 != 0:
                titolo_intero_1 = {
                    'tipo' : 'INTERO',
                    'serie_lettera' : re.findall('[a-zA-Z]+', fiscal_data.intero_ticket_serie_1),
                    'serie_numero' : re.findall('[0-9]+', fiscal_data.intero_ticket_serie_1),
                    'importo' : curr_event.price_full,
                    'da' : fiscal_data.intero_ticket_start_number_1,
                    'a' : fiscal_data.intero_ticket_end_number_1,
                }
                if titolo_intero_1['a']>= titolo_intero_1['da']:
                    titolo_intero_1['num'] = titolo_intero_1['a'] - titolo_intero_1['da'] + 1
                else:
                    titolo_intero_1['num'] = 0
            else:
                titolo_intero_1 = {
                    'tipo' : 'INTERO',
                    'serie_lettera' : 'nd',
                    'serie_numero' : 'nd',
                    'importo' : curr_event.price_full,
                    'da' : 0,
                    'a' : 0,
                    'num' : 0,
                }
            titoli.append(titolo_intero_1)

            if fiscal_data.intero_ticket_start_number_2 != 0:
                titolo_intero_2 = {
                    'tipo' : 'INTERO',
                    'serie_lettera' : re.findall('[a-zA-Z]+', fiscal_data.intero_ticket_serie_2),
                    'serie_numero' : re.findall('[0-9]+', fiscal_data.intero_ticket_serie_2),
                    'importo' : curr_event.price_full,
                    'da' : fiscal_data.intero_ticket_start_number_2,
                    'a' : fiscal_data.intero_ticket_end_number_2,
                }
                if titolo_intero_2['a']>= titolo_intero_2['da']:
                    titolo_intero_2['num'] = titolo_intero_2['a'] - titolo_intero_2['da'] + 1
                else:
                    titolo_intero_2['num'] = 0
                titoli.append(titolo_intero_2)

            if fiscal_data.ridotto_ticket_start_number_1 != 0:
                titolo_ridotto_1 = {
                    'tipo' : 'RIDOTTO',
                    'serie_lettera' : re.findall('[a-zA-Z]+', fiscal_data.ridotto_ticket_serie_1),
                    'serie_numero' : re.findall('[0-9]+', fiscal_data.ridotto_ticket_serie_1),
                    'importo' : curr_event.price_reduced,
                    'da' : fiscal_data.ridotto_ticket_start_number_1,
                    'a' : fiscal_data.ridotto_ticket_end_number_1,
                }
                if titolo_ridotto_1['a']>= titolo_ridotto_1['da']:
                    titolo_ridotto_1['num'] = titolo_ridotto_1['a'] - titolo_ridotto_1['da'] + 1
                else:
                    titolo_ridotto_1['num'] = 0
            else:
                titolo_ridotto_1 = {
                    'tipo' : 'RIDOTTO',
                    'serie_lettera' : 'nd',
                    'serie_numero' : 'nd',
                    'importo' : curr_event.price_reduced,
                    'da' : 0,
                    'a' : 0,
                    'num' : 0,
                }
            titoli.append(titolo_ridotto_1)


            if fiscal_data.ridotto_ticket_start_number_2 != 0:
                titolo_ridotto_2 = {
                    'tipo' : 'RIDOTTO',
                    'serie_lettera' : re.findall('[a-zA-Z]+', fiscal_data.ridotto_ticket_serie_2),
                    'serie_numero' : re.findall('[0-9]+', fiscal_data.ridotto_ticket_serie_2),
                    'importo' : curr_event.price_reduced,
                    'da' : fiscal_data.ridotto_ticket_start_number_2,
                    'a' : fiscal_data.ridotto_ticket_end_number_2,
                }
                if titolo_ridotto_2['a']>= titolo_ridotto_2['da']:
                    titolo_ridotto_2['num'] = titolo_ridotto_2['a'] - titolo_ridotto_2['da'] + 1
                else:
                    titolo_ridotto_2['num'] = 0
                titoli.append(titolo_ridotto_2)



            if fiscal_data.gratuito_ticket_start_number_1 != 0:
                titolo_gratuito_1 = {
                    'tipo' : 'GRATUITO',
                    'serie_lettera' : re.findall('[a-zA-Z]+', fiscal_data.gratuito_ticket_serie_1),
                    'serie_numero' : re.findall('[0-9]+', fiscal_data.gratuito_ticket_serie_1),
                    'importo' : 0.0,
                    'da' : fiscal_data.gratuito_ticket_start_number_1,
                    'a' : fiscal_data.gratuito_ticket_end_number_1,
                }
                if titolo_gratuito_1['a']>= titolo_gratuito_1['da']:
                    titolo_gratuito_1['num'] = titolo_gratuito_1['a'] - titolo_gratuito_1['da'] + 1
                else:
                    titolo_gratuito_1['num'] = 0
            else:
                titolo_gratuito_1 = {
                    'tipo' : 'GRATUITO',
                    'serie_lettera' : 'n.d.',
                    'serie_numero' : 'n.d.',
                    'importo' : 0.0,
                    'da' : 0,
                    'a' : 0,
                    'num' : 0,
                }
            titoli.append(titolo_gratuito_1)
            

            if fiscal_data.gratuito_ticket_start_number_2 != 0:
                titolo_gratuito_2 = {
                    'tipo' : 'GRATUITO',
                    'serie_lettera' : re.findall('[a-zA-Z]+', fiscal_data.gratuito_ticket_serie_2),
                    'serie_numero' : re.findall('[0-9]+', fiscal_data.gratuito_ticket_serie_2),
                    'importo' : 0.0,
                    'da' : fiscal_data.gratuito_ticket_start_number_2,
                    'a' : fiscal_data.gratuito_ticket_end_number_2,
                }
                if titolo_gratuito_2['a']>= titolo_gratuito_2['da']:
                    titolo_gratuito_2['num'] = titolo_gratuito_2['a'] - titolo_gratuito_2['da'] + 1
                else:
                    titolo_gratuito_2['num'] = 0
                titoli.append(titolo_gratuito_2)

            
            generate_mod2A: bool = True
            generate_mod566: bool = True
            generate_casher: bool = True

            if generate_mod2A: # MOD 2A Excel Generator Procedure
                mod2da_xls_filename = f'{evnt_date_code}_mod2DA_{evnt_show_slug}.xls'
                filename = os.path.join(os.getcwd(), save_path,mod2da_xls_filename)

                try: 
                    report = Report.objects.get(event_id = curr_event.pk, type ='SIAE_2DA')
                except:
                    report = Report()
                    report.type = 'SIAE_2DA'
                    report.event = curr_event
                    report.doc_path = mod2da_xls_filename
                    report.save()
                    try:
                        all_reports = Report.objects.filter(type = 'SIAE_2DA')
                        progress_number_old = 0
                        for report in all_reports:
                            if report.progress_number > progress_number_old:
                                progress_number_old = report.progress_number
                    except:
                        print('No previus SIAE_2DA') 
                    progress_number = progress_number_old + 1
                
                    report.progress_number = progress_number
                    report.save()

                
                wb = Workbook()
                ws = wb.active
                ws.title = 'mod 2DA'
                ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                ws.page_margins.left = 0.25
                ws.page_margins.right = 0.25
                ws.page_margins.top = 0.25
                ws.page_margins.bottom = 0.25

            
                ws.HeaderFooter.differentFirst = True
                ws.firstHeader.left.text = "Page &P of &N"
                ws.firstHeader.center.text = "MOD. 2DA"

                # setting the column size
                colnames = ('A','B','C','D','E','F','G','H','I','J')
                for col in colnames:
                    ws.column_dimensions[col].width = 10.95

                #Title merged cells
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
                tlc = ws['A1']
                tlc.value = "Prospetto di Liquidazione Diritti d'Autore e quote varie per manifestazioni musicali in genere, teatrali, di danza e letterarie"
                tlc.font = font['bold11']
                tlc.alignment = alignment['c_c']

                #SubTitle merged cells
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
                tlcs = ws['A2']
                tlcs.value = "(rif. Punto 14 delle condizioni generali del mod. 116)"
                tlcs.font = font['bold11']
                tlcs.alignment = alignment['c_c']

                #Titolo Riquadro Manifestazione
                try:
                    trm = ws['A3']
                    trm.value = "Manifestazione"
                    trm.font = font['bold11']
                    trm.alignment = alignment['l_b']

                    # Square Manifestazione
                    sq_man = ws['A4:J10']
                    make_border(sq_man,'double')
                    ws['A4'].value = 'Organizzatore:'
                    ws['A4'].font = font['normal']
                    ws['B4'].value = 'Laboratorio Teatrale di Cambiano APS'
                    ws['B4'].font = font['bold']

                    ws['F4'].value = 'Cod.Fiscale/Partita IVA:'
                    ws['F4'].font = font['normal']
                    ws['H4'].value = '09762270016'
                    ws['H4'].font = font['bold']

                    ws['A5'].value = 'Locale:'
                    ws['A5'].font = font['normal']
                    ws['B5'].value = 'Teatro Comunale di Cambiano'
                    ws['B5'].font = font['bold']

                    ws['F5'].value = 'Codice locale:'
                    ws['F5'].font = font['normal']
                    ws['H5'].value = '0050450366484'
                    ws['H5'].font = font['bold']

                    ws['A6'].value = 'Genere Manifestazione:'
                    ws['A6'].font = font['normal']
                    ws['C6'].value = '45 - Prosa Cabaret'
                    ws['C6'].font = font['bold']

                    ws['A7'].value = 'Titolo(1):'
                    ws['A7'].font = font['normal']
                    ws['B7'].value = curr_event.show.shw_title
                    ws['B7'].font = font['bold']

                    ws['F7'].value = 'Autore:'
                    ws['F7'].font = font['normal']
                    ws['H7'].value = curr_event.show.shw_author
                    ws['H7'].font = font['bold']

                    ws['A8'].value = 'Titolo(2):'
                    ws['A8'].font = font['normal']
                    ws['B8'].value = 'ND'
                    ws['B8'].font = font['bold']

                    ws['F8'].value = 'Autore:'
                    ws['F8'].font = font['normal']
                    ws['H8'].value = "ND"
                    ws['H8'].font = font['bold']

                    ws['A9'].value = 'Compagnia:'
                    ws['A9'].font = font['normal09']
                    ws['B9'].value = curr_event.show.shw_theater_company
                    ws['B9'].font = font['bold09']

                    ws['E9'].value = 'Direttore:'
                    ws['E9'].font = font['normal09']
                    ws['F9'].value = curr_event.show.shw_director
                    ws['F9'].font = font['bold09']
                
                    ws['H9'].value = 'Permesso SIAE:'
                    ws['H9'].font = font['normal09']
                    ws['I9'].value = "123456789123456"
                    ws['I9'].font = font['bold09']

                    ws['A10'].value = 'Data evento:'
                    ws['A10'].font = font['normal09']

                    ws['B10'].value = date_time.strftime('%d/%m/%Y')
                    ws['B10'].font = font['bold09']

                    ws['E10'].value = 'Ora evento:'
                    ws['E10'].font = font['normal09']
                    ws['F10'].value = date_time.strftime('%H:%M')
                    ws['F10'].font = font['bold09']
                
                    ws['H10'].value = 'GG SETT:'
                    ws['H10'].font = font['normal09']
                    ws['I10'].value = date_format(date_time,'l')
                    ws['I10'].font = font['bold09']


                except:
                    print("Riquadro Manifestazione - Possibili dati mancanti")


                #Titolo Riquadro Titoli d'accesso
                try:
                    begin_row = 11
                    trm = ws['A{}'.format(begin_row+0)]
                    trm.value = "Titoli di accesso emessi dai seguenti sistemi:"
                    trm.font = font['bold11']
                    trm.alignment = alignment['l_b']
                    ws['E{}'.format(begin_row+0)].value = 'Codice:'
                    ws['E{}'.format(begin_row+0)].font = font['bold']
                    make_underline(ws['F{}'.format(begin_row+0)],'thin')

                    ws['H{}'.format(begin_row+0)].value = 'Titolare:'
                    ws['H{}'.format(begin_row+0)].font = font['bold']
                    make_underline(ws['I{}:J{}'.format(begin_row+0,begin_row+0)],'thin')
        
                    # Square riquadratura
                    # sq_man = ws['A{}:J{}'.format(begin_row+1,begin_row+4)]
                    # make_border(sq_man,'double')
                    ws['A{}'.format(begin_row+1)].value = '(comprensivi della quota abbonamenti):'
                    ws['A{}'.format(begin_row+1)].font = font['normal']
                    ws['E{}'.format(begin_row+1)].value = 'Codice:'
                    ws['E{}'.format(begin_row+1)].font = font['bold']
                    make_underline(ws['F{}'.format(begin_row+1)],'thin')

                    ws['H{}'.format(begin_row+1)].value = 'Titolare:'
                    ws['H{}'.format(begin_row+1)].font = font['bold']
                    make_underline(ws['I{}:J{}'.format(begin_row+1,begin_row+1)],'thin')

                    ws['E{}'.format(begin_row+2)].value = 'Codice:'
                    ws['E{}'.format(begin_row+2)].font = font['bold']
                    make_underline(ws['F{}'.format(begin_row+2)],'thin')

                    ws['H{}'.format(begin_row+2)].value = 'Titolare:'
                    ws['H{}'.format(begin_row+2)].font = font['bold']
                    make_underline(ws['I{}:J{}'.format(begin_row+2,begin_row+2)],'thin')

                    ws['E{}'.format(begin_row+3)].value = 'Codice:'
                    ws['E{}'.format(begin_row+3)].font = font['bold']
                    make_underline(ws['F{}'.format(begin_row+3)],'thin')

                    ws['H{}'.format(begin_row+3)].value = 'Titolare:'
                    ws['H{}'.format(begin_row+3)].font = font['bold']
                    make_underline(ws['I{}:J{}'.format(begin_row+3,begin_row+3)],'thin')

                except:
                    print("Riquadro Titoli di accesso - Possibili dati mancanti")

                #Titolo Riquadro Presenze
                try:
                    begin_row = 16
                    row = begin_row 

                    trp = ws[f'A{row}']
                    trp.value = "Presenze:"
                    trp.font = font['bold11']
                    trp.alignment = alignment['l_b']
                    # Square riquadratura
                    sq_rp = ws[f'A{row+1}:C{row+5}']
                    make_border(sq_rp,'double')

                    row = begin_row + 1

                    ws[f'A{row}'].value = 'INGRESSO'
                    ws[f'A{row}'].font = font['normal']
                    make_underline(ws[f'A{row}'],'thin')
                    make_rightline(ws[f'A{row}'],'thin')
        
                    ws[f'B{row}'].value = 'Incasso lordo'
                    ws[f'B{row}'].font = font['normal09']
                    make_underline(ws[f'B{row}'],'thin')
                    make_rightline(ws[f'B{row}'],'thin')

                    ws[f'C{row}'].value = 'Incasso netto'
                    ws[f'C{row}'].font = font['normal09']
                    make_underline(ws[f'C{row}'],'thin')

                    row = begin_row + 2

                    ws[f'A{row}'].value = ingressi['globale'].count()
                    ws[f'A{row}'].font = font['normal']
                    make_underline(ws[f'A{row}'],'thin')
                    make_rightline(ws[f'A{row}'],'thin')
        
                    ws[f'B{row}'].value = introiti['globale']
                    ws[f'B{row}'].font = font['normal09']
                    ws[f'B{row}'].number_format  = '0.00'
                    make_underline(ws[f'B{row}'],'thin')
                    make_rightline(ws[f'B{row}'],'thin')

                    ws[f'C{row}'].value = introiti['globale']
                    ws[f'C{row}'].font = font['normal09']
                    ws[f'C{row}'].number_format  = '0.00'
                    make_underline(ws[f'C{row}'],'thin')

                    row = begin_row + 3

                    # un_iot 
                    ws.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=1)
                    tlc_un_iot = ws[f'A{row}']
                    tlc_un_iot.value = 'Ingressi omaggio'
                    tlc_un_iot.font = font['normal09']
                    tlc_un_iot.alignment = alignment['c_c']
                    make_underline(ws[f'A{row+1}'],'thin')
                    make_rightline(tlc_un_iot,'thin')
        
                    make_rightline(ws[f'B{row}'],'thin')
                    make_rightline(ws[f'B{row+1}'],'thin')

                    # un_maxprice 
                    ws.merge_cells(start_row=row, start_column=3, end_row=row+1, end_column=3)
                    tlc_un_maxprice = ws[f'C{row}']
                    tlc_un_maxprice.value = 'Biglietto di maggior costo in vendita (imp)'
                    tlc_un_maxprice.font = font['normal07']
                    tlc_un_maxprice.alignment = alignment['c_c']
                    make_underline(ws[f'C{row+1}'],'thin')

                    row = begin_row + 5

                    ws[f'A{row}'].value = ingressi['omaggio'].count()
                    ws[f'A{row}'].font = font['normal']
                    make_rightline(ws[f'A{row}'],'thin')
                    make_rightline(ws[f'B{row}'],'thin')
        
                    ws[f'C{row}'].value = curr_event.price_full
                    ws[f'C{row}'].font = font['normal09']
                    ws[f'C{row}'].number_format  = '0.00'
                except:
                    print("Riquadro Presenze - Possibili dati mancanti")


                #Titolo Riquadro Altri Proventi 
                try:
                    begin_row = 16
                    row = begin_row 

                    trap = ws[f'E{row}']
                    trap.value = "Altri proventi:"
                    trap.font = font['bold11']
                    trap.alignment = alignment['l_b']
                    # Square riquadratura
                    sq_rap = ws[f'E{row+1}:J{row+8}']
                    make_border(sq_rap,'double')

                    row = begin_row + 1

                    ws[f'E{row}'].value = 'Tipo provento'
                    ws[f'E{row}'].font = font['normal']
                    ws[f'E{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'E{row}'],'thin')
                    make_underline(ws[f'F{row}'],'thin')
                    make_rightline(ws[f'F{row}'],'thin')
                    ws.merge_cells(start_row=row, start_column=colnum('e'), end_row=row, end_column=colnum('f'))
        
                    ws[f'G{row}'].value = 'Lordo'
                    ws[f'G{row}'].font = font['normal09']
                    ws[f'G{row}'].alignment = alignment['c_c']
                    make_underline(ws[f'G{row}'],'thin')
                    make_underline(ws[f'H{row}'],'thin')
                    make_rightline(ws[f'H{row}'],'thin')
                    ws.merge_cells(start_row=row, start_column=colnum('g'), end_row=row, end_column=colnum('h'))

                    ws[f'I{row}'].value = 'Netto'
                    ws[f'I{row}'].font = font['normal09']
                    ws[f'I{row}'].alignment = alignment['c_c']
                    make_underline(ws[f'I{row}'],'thin')
                    make_underline(ws[f'J{row}'],'thin')
                    ws.merge_cells(start_row=row, start_column=colnum('i'), end_row=row, end_column=colnum('j'))

                    altri_proventi_types = [
                                            'Diritti di prevendita terzi',
                                            'Diritti di prevendita c/proprio',
                                            'Diritti di ripresa',
                                            'Contributi',
                                            '2e CONSUMAZIONI',
                                            'GUARDAROBA',
                                            ]

                    for offset, tipo in enumerate(altri_proventi_types):
                        row = begin_row + 2 + offset

                        ws[f'E{row}'].value = tipo
                        ws[f'E{row}'].font = font['normal']
                        ws[f'E{row}'].alignment = alignment['l_c']
                        make_rightline(ws[f'F{row}'],'thin')
                        ws.merge_cells(start_row=row, start_column=colnum('e'), end_row=row, end_column=colnum('f'))
            
                        ws[f'G{row}'].value = 0.00
                        ws[f'G{row}'].font = font['normal09']
                        ws[f'G{row}'].alignment = alignment['r_c']
                        ws[f'G{row}'].number_format  = '0.00'
                        make_rightline(ws[f'H{row}'],'thin')
                        ws.merge_cells(start_row=row, start_column=colnum('g'), end_row=row, end_column=colnum('h'))

                        ws[f'I{row}'].value = 0.00
                        ws[f'I{row}'].font = font['normal09']
                        ws[f'I{row}'].alignment = alignment['r_c']
                        ws[f'I{row}'].number_format  = '0.00'
                        ws.merge_cells(start_row=row, start_column=colnum('i'), end_row=row, end_column=colnum('j'))
                        if offset < (len(altri_proventi_types)-1):
                            make_underline(ws[f'E{row}'],'thin')
                            make_underline(ws[f'F{row}'],'thin')
                            make_underline(ws[f'G{row}'],'thin')
                            make_underline(ws[f'H{row}'],'thin')
                            make_underline(ws[f'I{row}'],'thin')
                            make_underline(ws[f'J{row}'],'thin')
                    del altri_proventi_types
                except:
                    print("Riquadro Altri proventi- Possibili dati mancanti")

                #Titolo Omaggio Diritti Autore
                try:
                    begin_row = 22
                    row = begin_row 

                    trap = ws[f'A{row}']
                    trap.value = "Omaggi soggetti a Diritto d'Autore"
                    trap.font = font['bold']
                    trap.alignment = alignment['l_b']
                    # Square riquadratura
                    sq_rap = ws[f'A{row+1}:C{row+2}']
                    make_border(sq_rap,'double')

                    row = begin_row + 1

                    ws[f'A{row}'].value = 'Num.biglietti'
                    ws[f'A{row}'].font = font['normal']
                    ws[f'A{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'A{row}'],'thin')
                    make_rightline(ws[f'A{row}'],'thin')
        
                    ws[f'B{row}'].value = 'Base calcolo'
                    ws[f'B{row}'].font = font['normal09']
                    ws[f'B{row}'].alignment = alignment['c_c']
                    make_underline(ws[f'B{row}'],'thin')
                    make_underline(ws[f'C{row}'],'thin')
                    ws.merge_cells(start_row=row, start_column=colnum('b'), end_row=row, end_column=colnum('c'))

                    row = begin_row + 2

                    ws[f'A{row}'].value = 0
                    ws[f'A{row}'].font = font['normal']
                    ws[f'A{row}'].alignment = alignment['r_c']
                    make_rightline(ws[f'A{row}'],'thin')
        
                    ws[f'B{row}'].value = 0.00
                    ws[f'B{row}'].font = font['normal09']
                    ws[f'B{row}'].alignment = alignment['r_c']
                    ws[f'B{row}'].number_format  = '0.00'
                    ws.merge_cells(start_row=row, start_column=colnum('b'), end_row=row, end_column=colnum('c'))


                    row = begin_row + 3

                    ws[f'A{row}'].value = "Si vedano in proposito le copie dei mod. C1 e C2 ovvero dei mod. SD1 e SD2 che si allegano e che costituiscono parte integrante del presente prospetto"
                    ws[f'A{row}'].font = font['bold09']
                    ws[f'A{row}'].alignment = alignment['l_c']

                except:
                    print("Riquadro Omaggio Diritti Autore- Possibili dati mancanti")


                #Titolo Riquadro Distinta di pagamento
                try:
                    begin_row = 27
                    row = begin_row 

                    DP = ws[f'A{row}']
                    DP.value = "Distinta di pagamento"
                    DP.font = font['bold11']
                    DP.alignment = alignment['l_b']
                    # Square riquadratura
                    sq_DP = ws[f'A{row+1}:J{row+7}']
                    make_border(sq_DP,'double')

                    DP_types = [
                                ' ',
                                'Diritti di esecuzione musicale (ASSOMUSICA)',
                                'Diritti di esecuzione musicale',
                                'Diritti di esecuzione musicale',
                                'Quota associativa FIPE',
                                'Quota FIPE/ARAG',
                                'Quota ASSOMUSICA (0,5%)',
                                ]

                    for offset, tipo in enumerate(DP_types):
                        row = begin_row + 1 + offset

                        ws[f'A{row}'].value = tipo
                        ws[f'A{row}'].font = font['normal']
                        ws[f'A{row}'].alignment = alignment['l_c']
                        make_rightline(ws[f'D{row}'],'thin')
                        ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row, end_column=colnum('d'))

                        if offset==0:
                            ws[f'E{row}'].value = 'Descrizione'
                        else:
                            ws[f'E{row}'].value = ' '
                        ws[f'E{row}'].font = font['normal']
                        ws[f'E{row}'].alignment = alignment['c_c']
                        make_rightline(ws[f'F{row}'],'thin')
                        ws.merge_cells(start_row=row, start_column=colnum('e'), end_row=row, end_column=colnum('f'))

                        if offset==0:
                            ws[f'G{row}'].value = 'Importo'
                        else:
                            ws[f'G{row}'].value = 0.00
                        ws[f'G{row}'].font = font['normal09']
                        ws[f'G{row}'].alignment = alignment['r_c']
                        ws[f'G{row}'].number_format  = '0.00'
                        make_rightline(ws[f'G{row}'],'thin')

                        if offset==0:
                            ws[f'H{row}'].value = 'Descrizione'
                        else:
                            ws[f'H{row}'].value = ' '
                        ws[f'H{row}'].font = font['normal']
                        ws[f'H{row}'].alignment = alignment['c_c']
                        make_rightline(ws[f'I{row}'],'thin')
                        ws.merge_cells(start_row=row, start_column=colnum('h'), end_row=row, end_column=colnum('i'))

                        if offset==0:
                            ws[f'J{row}'].value = 'Importo'
                        else:
                            ws[f'J{row}'].value = 0.00
                        ws[f'J{row}'].font = font['normal09']
                        ws[f'J{row}'].alignment = alignment['r_c']
                        ws[f'J{row}'].number_format  = '0.00'

                        if offset < (len(DP_types)-1):
                            make_underline(ws[f'A{row}'],'thin')
                            make_underline(ws[f'B{row}'],'thin')
                            make_underline(ws[f'C{row}'],'thin')
                            make_underline(ws[f'D{row}'],'thin')
                            make_underline(ws[f'E{row}'],'thin')
                            make_underline(ws[f'F{row}'],'thin')
                            make_underline(ws[f'G{row}'],'thin')
                            make_underline(ws[f'H{row}'],'thin')
                            make_underline(ws[f'I{row}'],'thin')
                            make_underline(ws[f'J{row}'],'thin')
                    del DP_types
                except:
                    print("Riquadro Distinta di pagamento- Possibili dati mancanti")

                try: # Data e firma
                    begin_row = 35

                    row = begin_row
                    ws[f'G{row}'].value = 'Gli importi sono espressi in Euro'
                    ws[f'G{row}'].font = font['normal09']
                    ws[f'G{row}'].alignment = alignment['l_c']



                    row = begin_row + 1 

                    ws[f'A{row}'].value = 'Data'
                    ws[f'A{row}'].font = font['bold09']
                    ws[f'A{row}'].alignment = alignment['l_c']

                    make_underline(ws[f'B{row}'],'thin')
                    make_underline(ws[f'C{row}'],'thin')
                    make_underline(ws[f'D{row}'],'thin')

                    ws[f'F{row}'].value = 'Firma organizzatore'
                    ws[f'F{row}'].font = font['bold09']
                    ws[f'F{row}'].alignment = alignment['r_c']


                    make_underline(ws[f'G{row}'],'thin')
                    make_underline(ws[f'H{row}'],'thin')
                    make_underline(ws[f'I{row}'],'thin')
                except:
                    print("Riquadro Data e firma- Possibili dati mancanti")

                ws.row_dimensions[15].height = 2
                ws.row_dimensions[26].height = 2
                ws.row_dimensions[36].height = 30


                wb.save(filename=filename)
                report.updated_at = datetime.now()
                report.save()

            if generate_mod566: # MOD 566 Excel Generator Procedure
                mod566_xls_filename = f'{evnt_date_code}_mod566_{evnt_show_slug}.xls'
                filename = os.path.join(os.getcwd(), save_path,mod566_xls_filename)   

                try: 
                    report = Report.objects.get(event_id = curr_event.pk, type ='SIAE_566')
                except:
                    report = Report()
                    report.type = 'SIAE_566'
                    report.event = curr_event
                    report.doc_path = mod566_xls_filename
                    report.save()
                    try:
                        all_reports = Report.objects.filter(type = 'SIAE_566')
                        progress_number_old = 0
                        for report in all_reports:
                            if report.progress_number > progress_number_old:
                                progress_number_old = report.progress_number
                    except:
                        print('No previus SIAE_566') 
                    progress_number = progress_number_old + 1
                
                    report.progress_number = progress_number
                    report.save()

                wb = Workbook()
                ws = wb.active
                ws.title = 'mod SD1'
                ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                ws.page_margins.left = 0.25
                ws.page_margins.right = 0.25
                ws.page_margins.top = 0.25
                ws.page_margins.bottom = 0.25
                
                ws.HeaderFooter.differentFirst = True
                ws.firstHeader.left.text = "Page &P of &N"
                ws.firstHeader.center.text = "MOD. 566"

                # setting the column size
                colnames = ('A','B','C','D','E','F','G','H','I','J')
                for col in colnames:
                    ws.column_dimensions[col].width = 7.7
                img = Image(os.path.join(os.getcwd(), save_path,'AgenziaEntrateLogo.png'))
                img.height = 442 /6 # insert image height in pixels as float or int (e.g. 305.5)
                img.width= 1874 / 6 # insert image width in pixels as float or int (e.g. 405.8)
                img.anchor = 'A1' # where you want image to be anchored/start from
                ws.add_image(img) # adding in the image

                # Headers
                begin_row = 1
                row = begin_row + 0
                ws[f'J{row}'].value = 'Mod.SD/1'
                ws[f'J{row}'].font = font['bold']
                ws[f'J{row}'].alignment = alignment['r_c']

                begin_row = 3
                row = begin_row + 0
                ws.merge_cells(start_row=row, start_column=colnum('f'), end_row=row, end_column=colnum('g'))
                ws[f'F{row}'].value = 'N. Progressivo'
                ws[f'F{row}'].font = font['normal09']
                ws[f'F{row}'].alignment = alignment['r_c']

                ws[f'H{row}'].value = report.progress_number
                ws[f'H{row}'].font = font['bold09']
                ws[f'H{row}'].alignment = alignment['r_c']
                ws[f'H{row}'].number_format = '00000'
                make_border(ws[f'H{row}'],'thin')

                ws[f'I{row}'].value = 'Foglio n.'
                ws[f'I{row}'].font = font['normal09']
                ws[f'I{row}'].alignment = alignment['r_c']
                ws[f'J{row}'].value = 1
                ws[f'J{row}'].font = font['bold09']
                ws[f'J{row}'].alignment = alignment['r_c']
                ws[f'J{row}'].number_format = '00'
                make_border(ws[f'J{row}'],'thin')



                begin_row = 6
                row = begin_row + 0
                ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row, end_column=colnum('j'))
                ws[f'A{row}'].value = "PROSPETTO RIEPILOGATIVO DEI TITOLI DI INGRESSO RILASCIATI"
                ws[f'A{row}'].font = font['bold13']
                ws[f'A{row}'].alignment = alignment['c_c']
                row = begin_row + 1
                ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row, end_column=colnum('j'))
                ws[f'A{row}'].value = "PER CIASCUNA MANIFESTAZIONE art.5, comma 3, DPR 13 marzo 2002, n.69"
                ws[f'A{row}'].font = font['bold13']
                ws[f'A{row}'].alignment = alignment['c_c']

                # Titolo Riquadro Manifestazione
                try:
                    begin_row = 9
                    row = begin_row + 0

                    ws[f'C{row}'].value = 'Giorno'
                    ws[f'C{row}'].font = font['normal']
                    ws[f'C{row}'].alignment = alignment['c_c']
                    ws[f'D{row}'].value = 'Mese'
                    ws[f'D{row}'].font = font['normal']
                    ws[f'D{row}'].alignment = alignment['c_c']
                    ws[f'E{row}'].value = 'Anno'
                    ws[f'E{row}'].font = font['normal']
                    ws[f'E{row}'].alignment = alignment['c_c']
                    
                    ws[f'F{row}'].value = 'Titolo:'
                    ws[f'F{row}'].font = font['normal09']
                    ws[f'F{row}'].alignment = alignment['r_c']

                    ws.merge_cells(start_row=row, start_column=colnum('g'), end_row=row, end_column=colnum('j'))
                    ws[f'G{row}'].value = f'{curr_event.show.shw_title}'
                    ws[f'G{row}'].font = font['bold09']
                    ws[f'G{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'G{row}'],'thin')
                    make_underline(ws[f'I{row}'],'thin')
                    make_underline(ws[f'J{row}'],'thin')

                    row = begin_row + 1

                    ws[f'A{row}'].value = 'Manifestazione:'
                    ws[f'A{row}'].font = font['normal09']
                    ws[f'A{row}'].alignment = alignment['l_c']

                    ws[f'C{row}'].value = date_time.strftime('%d')
                    ws[f'C{row}'].font = font['bold09']
                    ws[f'C{row}'].alignment = alignment['c_c']
                    make_border(ws[f'C{row}'],'thin')
                    ws[f'D{row}'].value = date_time.strftime('%m')
                    ws[f'D{row}'].font = font['bold09']
                    ws[f'D{row}'].alignment = alignment['c_c']
                    make_border(ws[f'D{row}'],'thin')
                    ws[f'E{row}'].value = date_time.strftime('%Y')
                    ws[f'E{row}'].font = font['bold09']
                    ws[f'E{row}'].alignment = alignment['c_c']
                    make_border(ws[f'E{row}'],'thin')

                    ws[f'F{row}'].value = 'Ora'
                    ws[f'F{row}'].font = font['normal09']
                    ws[f'F{row}'].alignment = alignment['r_c']

                    ws[f'G{row}'].value = date_time.strftime('%H:%M')
                    ws[f'G{row}'].font = font['bold09']
                    ws[f'G{row}'].alignment = alignment['c_c']
                    make_border(ws[f'G{row}'],'thin')

                    ws[f'H{row}'].value = 'Tipo: '
                    ws[f'H{row}'].font = font['normal09']
                    ws[f'H{row}'].alignment = alignment['r_c']

                    ws[f'I{row}'].value = f'{curr_event.show.siaetype}'
                    ws[f'I{row}'].font = font['normal09']
                    ws[f'I{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'I{row}'],'thin')
                    make_underline(ws[f'J{row}'],'thin')

                    row = begin_row + 2
                    ws.row_dimensions[row].height = 2
                except:
                    print("Riquadro Manifestazione - Possibili dati mancanti")

                # Titolo Riquadro ANAGRAFICA
                try:
                    row = begin_row + 3
                    ws[f'A{row}'].value = 'Società/Associazione'
                    ws[f'A{row}'].font = font['normal09']
                    ws[f'A{row}'].alignment = alignment['l_c']

                    ws[f'C{row}'].value = "LABORATORIO TEATRALE DI CAMBIANO APS"
                    ws[f'C{row}'].font = font['bold09']
                    ws[f'C{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'C{row}'],'thin')
                    make_underline(ws[f'D{row}'],'thin')
                    make_underline(ws[f'E{row}'],'thin')
                    make_underline(ws[f'F{row}'],'thin')

                    ws.merge_cells(start_row=row, start_column=colnum('g'), end_row=row, end_column=colnum('h'))
                    ws[f'G{row}'].value = 'Codice fiscale:'
                    ws[f'G{row}'].font = font['normal09']
                    ws[f'G{row}'].alignment = alignment['r_c']

                    ws.merge_cells(start_row=row, start_column=colnum('i'), end_row=row, end_column=colnum('j'))
                    ws[f'I{row}'].value = "09762270016"
                    ws[f'I{row}'].font = font['bold09']
                    ws[f'I{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'I{row}'],'thin')
                    make_underline(ws[f'J{row}'],'thin')

                    row = begin_row + 4
                    ws.row_dimensions[row].height = 2

                    row = begin_row + 5
                    ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row, end_column=colnum('c'))
                    ws[f'A{row}'].value = 'Sede legale: via/piazza:'
                    ws[f'A{row}'].font = font['normal09']
                    ws[f'A{row}'].alignment = alignment['r_c']

                    ws.merge_cells(start_row=row, start_column=colnum('d'), end_row=row, end_column=colnum('f'))
                    ws[f'D{row}'].value = "via LAME"
                    ws[f'D{row}'].font = font['bold09']
                    ws[f'D{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'D{row}'],'thin')
                    make_underline(ws[f'E{row}'],'thin')
                    make_underline(ws[f'F{row}'],'thin')

                    ws[f'G{row}'].value = 'Num.'
                    ws[f'G{row}'].font = font['normal09']
                    ws[f'G{row}'].alignment = alignment['r_c']

                    ws[f'H{row}'].value = "4"
                    ws[f'H{row}'].font = font['bold09']
                    ws[f'H{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'H{row}'],'thin')

                    ws[f'I{row}'].value = 'CAP'
                    ws[f'I{row}'].font = font['normal09']
                    ws[f'I{row}'].alignment = alignment['r_c']

                    ws[f'J{row}'].value = "10020"
                    ws[f'J{row}'].font = font['bold09']
                    ws[f'J{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'J{row}'],'thin')

                    row = begin_row + 6
                    ws.row_dimensions[row].height = 2

                    row = begin_row + 7
                    ws[f'A{row}'].value = 'Comune:'
                    ws[f'A{row}'].font = font['normal09']
                    ws[f'A{row}'].alignment = alignment['r_c']

                    ws.merge_cells(start_row=row, start_column=colnum('b'), end_row=row, end_column=colnum('f'))
                    ws[f'B{row}'].value = "Cambiano"
                    ws[f'B{row}'].font = font['bold09']
                    ws[f'B{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'B{row}'],'thin')
                    make_underline(ws[f'C{row}'],'thin')
                    make_underline(ws[f'D{row}'],'thin')
                    make_underline(ws[f'E{row}'],'thin')
                    make_underline(ws[f'F{row}'],'thin')

                    ws[f'I{row}'].value = 'Prov.'
                    ws[f'I{row}'].font = font['normal09']
                    ws[f'I{row}'].alignment = alignment['r_c']

                    ws[f'J{row}'].value = "TO"
                    ws[f'J{row}'].font = font['bold09']
                    ws[f'J{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'J{row}'],'thin')
                except:
                    print("Riquadro ANAGRAFICA - Possibili dati mancanti")


                # Titolo Riquadro Anno Sociale
                try:
                    begin_row = 17
                    row = begin_row + 0

                    ws[f'D{row}'].value = 'Giorno'
                    ws[f'D{row}'].font = font['normal']
                    ws[f'D{row}'].alignment = alignment['c_c']
                    ws[f'E{row}'].value = 'Mese'
                    ws[f'E{row}'].font = font['normal']
                    ws[f'E{row}'].alignment = alignment['c_c']
                    ws[f'F{row}'].value = 'Anno'
                    ws[f'F{row}'].font = font['normal']
                    ws[f'F{row}'].alignment = alignment['c_c']


                    ws[f'H{row}'].value = 'Giorno'
                    ws[f'H{row}'].font = font['normal']
                    ws[f'H{row}'].alignment = alignment['c_c']
                    ws[f'I{row}'].value = 'Mese'
                    ws[f'I{row}'].font = font['normal']
                    ws[f'I{row}'].alignment = alignment['c_c']
                    ws[f'J{row}'].value = 'Anno'
                    ws[f'J{row}'].font = font['normal']
                    ws[f'J{row}'].alignment = alignment['c_c']

                    row = begin_row + 1

                    ws[f'A{row}'].value = 'Esercizio sociale:     data inizio'
                    ws[f'A{row}'].font = font['normal09']
                    ws[f'A{row}'].alignment = alignment['l_c']

                    ws[f'D{row}'].value = esercizio_sociale_inizio.strftime('%d')
                    ws[f'D{row}'].font = font['bold09']
                    ws[f'D{row}'].alignment = alignment['c_c']
                    make_border(ws[f'D{row}'],'thin')
                    ws[f'E{row}'].value = esercizio_sociale_inizio.strftime('%m')
                    ws[f'E{row}'].font = font['bold09']
                    ws[f'E{row}'].alignment = alignment['c_c']
                    make_border(ws[f'E{row}'],'thin')
                    ws[f'F{row}'].value = esercizio_sociale_inizio.strftime('%Y')
                    ws[f'F{row}'].font = font['bold09']
                    ws[f'F{row}'].alignment = alignment['c_c']
                    make_border(ws[f'F{row}'],'thin')

                    ws[f'G{row}'].value = 'data fine'
                    ws[f'G{row}'].font = font['normal09']
                    ws[f'G{row}'].alignment = alignment['l_c']

                    ws[f'H{row}'].value = esercizio_sociale_fine.strftime('%d')
                    ws[f'H{row}'].font = font['bold09']
                    ws[f'H{row}'].alignment = alignment['c_c']
                    make_border(ws[f'H{row}'],'thin')
                    ws[f'I{row}'].value = esercizio_sociale_fine.strftime('%m')
                    ws[f'I{row}'].font = font['bold09']
                    ws[f'I{row}'].alignment = alignment['c_c']
                    make_border(ws[f'I{row}'],'thin')
                    ws[f'J{row}'].value = esercizio_sociale_fine.strftime('%Y')
                    ws[f'J{row}'].font = font['bold09']
                    ws[f'J{row}'].alignment = alignment['c_c']
                    make_border(ws[f'J{row}'],'thin')

                    row = begin_row + 2
                    ws.row_dimensions[row].height = 2
                except:
                    print("Riquadro Anno Sociale - Possibili dati mancanti")

                # Titolo Riquadro IMPIANTO
                try:
                    begin_row = 20
                    row = begin_row + 0
                    ws[f'A{row}'].value = 'Impianto:'
                    ws[f'A{row}'].font = font['normal09']
                    ws[f'A{row}'].alignment = alignment['l_c']

                    ws.merge_cells(start_row=row, start_column=colnum('c'), end_row=row, end_column=colnum('j'))
                    ws[f'C{row}'].value = "TEATRO COMUNALE DI CAMBIANO"
                    ws[f'C{row}'].font = font['bold09']
                    ws[f'C{row}'].alignment = alignment['c_c']
                    make_underline(ws[f'C{row}'],'thin')
                    make_underline(ws[f'D{row}'],'thin')
                    make_underline(ws[f'E{row}'],'thin')
                    make_underline(ws[f'F{row}'],'thin')
                    make_underline(ws[f'G{row}'],'thin')
                    make_underline(ws[f'H{row}'],'thin')
                    make_underline(ws[f'I{row}'],'thin')
                    make_underline(ws[f'J{row}'],'thin')
                    
                    row = begin_row + 1
                    ws[f'D{row}'].value = 'numero'
                    ws[f'D{row}'].font = font['normal']
                    ws[f'D{row}'].alignment = alignment['c_c']
                    ws[f'G{row}'].value = 'numero'
                    ws[f'G{row}'].font = font['normal']
                    ws[f'G{row}'].alignment = alignment['c_c']
                    ws[f'J{row}'].value = 'numero'
                    ws[f'J{row}'].font = font['normal']
                    ws[f'J{row}'].alignment = alignment['c_c']

                    row = begin_row + 2

                    ws[f'A{row}'].value = 'Capienza:'
                    ws[f'A{row}'].font = font['normal09']
                    ws[f'A{row}'].alignment = alignment['l_c']

                    ws.merge_cells(start_row=row, start_column=colnum('b'), end_row=row, end_column=colnum('c'))
                    ws[f'B{row}'].value = 'PLATEA'
                    ws[f'B{row}'].font = font['bold09']
                    ws[f'B{row}'].alignment = alignment['r_c']


                    ws[f'D{row}'].value = '234'
                    ws[f'D{row}'].font = font['bold09']
                    ws[f'D{row}'].alignment = alignment['r_c']
                    make_border(ws[f'D{row}'],'thin')

                    ws.merge_cells(start_row=row, start_column=colnum('e'), end_row=row, end_column=colnum('f'))

                    ws[f'G{row}'].value = ' '
                    ws[f'G{row}'].font = font['bold09']
                    ws[f'G{row}'].alignment = alignment['c_c']
                    make_border(ws[f'G{row}'],'thin')

                    ws.merge_cells(start_row=row, start_column=colnum('h'), end_row=row, end_column=colnum('i'))

                    ws[f'J{row}'].value = ' '
                    ws[f'J{row}'].font = font['bold09']
                    ws[f'J{row}'].alignment = alignment['c_c']
                    make_border(ws[f'J{row}'],'thin')

                    row = begin_row + 3
                    ws[f'D{row}'].value = 'numero'
                    ws[f'D{row}'].font = font['normal']
                    ws[f'D{row}'].alignment = alignment['c_c']
                    ws[f'G{row}'].value = 'numero'
                    ws[f'G{row}'].font = font['normal']
                    ws[f'G{row}'].alignment = alignment['c_c']
                    ws[f'J{row}'].value = 'numero'
                    ws[f'J{row}'].font = font['normal']
                    ws[f'J{row}'].alignment = alignment['c_c']

                    row = begin_row + 4


                    ws.merge_cells(start_row=row, start_column=colnum('b'), end_row=row, end_column=colnum('c'))
                    ws[f'B{row}'].value = ' '
                    ws[f'B{row}'].font = font['bold09']
                    ws[f'B{row}'].alignment = alignment['r_c']


                    ws[f'D{row}'].value = ' '
                    ws[f'D{row}'].font = font['bold09']
                    ws[f'D{row}'].alignment = alignment['r_c']
                    make_border(ws[f'D{row}'],'thin')

                    ws.merge_cells(start_row=row, start_column=colnum('e'), end_row=row, end_column=colnum('f'))

                    ws[f'G{row}'].value = ' '
                    ws[f'G{row}'].font = font['bold09']
                    ws[f'G{row}'].alignment = alignment['c_c']
                    make_border(ws[f'G{row}'],'thin')

                    ws.merge_cells(start_row=row, start_column=colnum('h'), end_row=row, end_column=colnum('i'))

                    ws[f'J{row}'].value = ' '
                    ws[f'J{row}'].font = font['bold09']
                    ws[f'J{row}'].alignment = alignment['c_c']
                    make_border(ws[f'J{row}'],'thin')

                    row = begin_row + 5
                    ws.row_dimensions[row].height = 2

                    row = begin_row + 6
                    ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row, end_column=colnum('c'))
                    ws[f'A{row}'].value = 'via/piazza:'
                    ws[f'A{row}'].font = font['normal09']
                    ws[f'A{row}'].alignment = alignment['r_c']

                    ws.merge_cells(start_row=row, start_column=colnum('d'), end_row=row, end_column=colnum('f'))
                    ws[f'D{row}'].value = "via LAME"
                    ws[f'D{row}'].font = font['bold09']
                    ws[f'D{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'D{row}'],'thin')
                    make_underline(ws[f'E{row}'],'thin')
                    make_underline(ws[f'F{row}'],'thin')

                    ws[f'G{row}'].value = 'Num.'
                    ws[f'G{row}'].font = font['normal09']
                    ws[f'G{row}'].alignment = alignment['r_c']

                    ws[f'H{row}'].value = "4"
                    ws[f'H{row}'].font = font['bold09']
                    ws[f'H{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'H{row}'],'thin')

                    ws[f'I{row}'].value = 'CAP'
                    ws[f'I{row}'].font = font['normal09']
                    ws[f'I{row}'].alignment = alignment['r_c']

                    ws[f'J{row}'].value = "10020"
                    ws[f'J{row}'].font = font['bold09']
                    ws[f'J{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'J{row}'],'thin')

                    row = begin_row + 7
                    ws.row_dimensions[row].height = 2

                    row = begin_row + 8
                    ws[f'A{row}'].value = 'Comune:'
                    ws[f'A{row}'].font = font['normal09']
                    ws[f'A{row}'].alignment = alignment['r_c']

                    ws.merge_cells(start_row=row, start_column=colnum('b'), end_row=row, end_column=colnum('f'))
                    ws[f'B{row}'].value = "Cambiano"
                    ws[f'B{row}'].font = font['bold09']
                    ws[f'B{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'B{row}'],'thin')
                    make_underline(ws[f'C{row}'],'thin')
                    make_underline(ws[f'D{row}'],'thin')
                    make_underline(ws[f'E{row}'],'thin')
                    make_underline(ws[f'F{row}'],'thin')

                    ws[f'I{row}'].value = 'Prov.'
                    ws[f'I{row}'].font = font['normal09']
                    ws[f'I{row}'].alignment = alignment['r_c']

                    ws[f'J{row}'].value = "TO"
                    ws[f'J{row}'].font = font['bold09']
                    ws[f'J{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'J{row}'],'thin')

                    row = begin_row + 9
                    ws.row_dimensions[row].height = 2
                except:
                    print("Riquadro IMPIANTO - Possibili dati mancanti")



                # Titolo Riquadro  Rappresentante Legale
                try:
                    begin_row = 30
                    row = begin_row + 0
                    ws[f'A{row}'].value = 'Rappresentante Legale:'
                    ws[f'A{row}'].font = font['bold09']
                    ws[f'A{row}'].alignment = alignment['l_c']

                    row = begin_row + 1
                    ws.row_dimensions[row].height = 2
                
                    row = begin_row + 2

                    ws[f'A{row}'].value = 'Cognome'
                    ws[f'A{row}'].font = font['normal09']
                    ws[f'A{row}'].alignment = alignment['r_c']

                    ws.merge_cells(start_row=row, start_column=colnum('b'), end_row=row, end_column=colnum('c'))
                    ws[f'B{row}'].value = 'Miletto'
                    ws[f'B{row}'].font = font['bold09']
                    ws[f'B{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'B{row}'],'thin')
                    make_underline(ws[f'C{row}'],'thin')
                    make_underline(ws[f'D{row}'],'thin')


                    ws[f'E{row}'].value = 'Nome'
                    ws[f'E{row}'].font = font['normal09']
                    ws[f'E{row}'].alignment = alignment['r_c']

                    ws.merge_cells(start_row=row, start_column=colnum('f'), end_row=row, end_column=colnum('g'))

                    ws[f'F{row}'].value = 'Giuseppe'
                    ws[f'F{row}'].font = font['bold09']
                    ws[f'F{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'F{row}'],'thin')
                    make_underline(ws[f'G{row}'],'thin')


                    ws[f'H{row}'].value = 'Cod Fiscale'
                    ws[f'H{row}'].font = font['normal09']
                    ws[f'H{row}'].alignment = alignment['r_c']
                    
                    ws.merge_cells(start_row=row, start_column=colnum('i'), end_row=row, end_column=colnum('j'))
                    ws[f'I{row}'].value = 'MLTGPP64E12F889Y'
                    ws[f'I{row}'].font = font['bold09']
                    ws[f'I{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'I{row}'],'thin')
                    make_underline(ws[f'J{row}'],'thin')

                    row = begin_row + 3
                    ws.row_dimensions[row].height = 2

                    row = begin_row + 4
                    ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row, end_column=colnum('e'))
                    ws[f'A{row}'].value = 'Residenza anagrafica o Domicilio fiscale: Via/piazza'
                    ws[f'A{row}'].font = font['normal09']
                    ws[f'A{row}'].alignment = alignment['r_c']

                    ws.merge_cells(start_row=row, start_column=colnum('f'), end_row=row, end_column=colnum('j'))
                    ws[f'F{row}'].value = "via Campi Rotondi 35"
                    ws[f'F{row}'].font = font['bold09']
                    ws[f'F{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'F{row}'],'thin')
                    make_underline(ws[f'G{row}'],'thin')
                    make_underline(ws[f'H{row}'],'thin')
                    make_underline(ws[f'I{row}'],'thin')
                    make_underline(ws[f'J{row}'],'thin')

                    row = begin_row + 5
                    ws.row_dimensions[row].height = 2

                    row = begin_row + 6
                    ws[f'A{row}'].value = 'Comune:'
                    ws[f'A{row}'].font = font['normal09']
                    ws[f'A{row}'].alignment = alignment['r_c']

                    ws.merge_cells(start_row=row, start_column=colnum('b'), end_row=row, end_column=colnum('f'))
                    ws[f'B{row}'].value = "Cambiano"
                    ws[f'B{row}'].font = font['bold09']
                    ws[f'B{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'B{row}'],'thin')
                    make_underline(ws[f'C{row}'],'thin')
                    make_underline(ws[f'D{row}'],'thin')
                    make_underline(ws[f'E{row}'],'thin')
                    make_underline(ws[f'F{row}'],'thin')

                    ws[f'I{row}'].value = 'Prov.'
                    ws[f'I{row}'].font = font['normal09']
                    ws[f'I{row}'].alignment = alignment['r_c']

                    ws[f'J{row}'].value = "TO"
                    ws[f'J{row}'].font = font['bold09']
                    ws[f'J{row}'].alignment = alignment['l_c']
                    make_underline(ws[f'J{row}'],'thin')

                    row = begin_row + 7
                    ws.row_dimensions[row].height = 35
                except:
                    print("Riquadro Rappresentante Legale - Possibili dati mancanti") 



                # Titolo Riquadro  Titoli Ingressi

                try:
                    begin_row = 38
                
                    row = begin_row-1
                    for letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'):
                        make_underline(ws[f'{letter}{row}'], 'thin')


                
                    row = begin_row + 0

                    ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row, end_column=colnum('d'))
                    ws[f'A{row}'].value = 'TIPO TITOLO'
                    ws[f'A{row}'].font = font['normal']
                    ws[f'A{row}'].alignment = alignment['c_c']


                    ws.merge_cells(start_row=row, start_column=colnum('e'), end_row=row, end_column=colnum('f'))
                    ws[f'E{row}'].value = 'SERIE'
                    ws[f'E{row}'].font = font['normal']
                    ws[f'E{row}'].alignment = alignment['c_c']

                    ws[f'G{row}'].value = 'importo'
                    ws[f'G{row}'].font = font['normal09']
                    ws[f'G{row}'].alignment = alignment['c_c']

                    ws[f'H{row}'].value = 'Dal'
                    ws[f'H{row}'].font = font['normal09']
                    ws[f'H{row}'].alignment = alignment['c_c']

                    ws[f'I{row}'].value = 'Al'
                    ws[f'I{row}'].font = font['normal09']
                    ws[f'I{row}'].alignment = alignment['c_c']


                    ws[f'J{row}'].value = 'Quantità'
                    ws[f'J{row}'].font = font['normal09']
                    ws[f'J{row}'].alignment = alignment['c_c']

                    make_leftline(ws[f'A{row}'],'thin')
                    make_leftline(ws[f'E{row}'],'thin')
                    make_leftline(ws[f'G{row}'],'thin')
                    make_leftline(ws[f'H{row}'],'thin')
                    make_leftline(ws[f'I{row}'],'thin')
                    make_leftline(ws[f'J{row}'],'thin')
                    make_rightline(ws[f'J{row}'],'thin')

                    row = begin_row + 1

                    ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row, end_column=colnum('d'))
                    ws[f'A{row}'].value = '(come da denominazione stampata)'
                    ws[f'A{row}'].font = font['normal']
                    ws[f'A{row}'].alignment = alignment['c_c']

                    ws[f'E{row}'].value = 'lettera'
                    ws[f'E{row}'].font = font['normal09']
                    ws[f'E{row}'].alignment = alignment['c_c']

                    ws[f'F{row}'].value = 'numero'
                    ws[f'F{row}'].font = font['normal09']
                    ws[f'F{row}'].alignment = alignment['c_c']

                    ws[f'G{row}'].value = 'lordo unitario'
                    ws[f'G{row}'].font = font['normal07']
                    ws[f'G{row}'].alignment = alignment['c_c']

                    ws[f'H{row}'].value = 'numero'
                    ws[f'H{row}'].font = font['normal09']
                    ws[f'H{row}'].alignment = alignment['c_c']

                    ws[f'I{row}'].value = 'numero'
                    ws[f'I{row}'].font = font['normal09']
                    ws[f'I{row}'].alignment = alignment['c_c']


                    ws[f'J{row}'].value = '(col.6-col.5 +1)'
                    ws[f'J{row}'].font = font['normal07']
                    ws[f'J{row}'].alignment = alignment['c_c']

                    make_leftline(ws[f'A{row}'],'thin')
                    make_leftline(ws[f'E{row}'],'thin')
                    make_leftline(ws[f'G{row}'],'thin')
                    make_leftline(ws[f'H{row}'],'thin')
                    make_leftline(ws[f'I{row}'],'thin')
                    make_leftline(ws[f'J{row}'],'thin')
                    make_rightline(ws[f'J{row}'],'thin')

                    for letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'):
                        make_underline(ws[f'{letter}{row}'], 'thin')

                    row = begin_row + 2

                    ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row, end_column=colnum('d'))
                    ws[f'A{row}'].value = '1'
                    ws[f'A{row}'].font = font['normal']
                    ws[f'A{row}'].alignment = alignment['c_c']

                    ws[f'E{row}'].value = '2'
                    ws[f'E{row}'].font = font['normal']
                    ws[f'E{row}'].alignment = alignment['c_c']

                    ws[f'F{row}'].value = '3'
                    ws[f'F{row}'].font = font['normal']
                    ws[f'F{row}'].alignment = alignment['c_c']

                    ws[f'G{row}'].value = '4'
                    ws[f'G{row}'].font = font['normal']
                    ws[f'G{row}'].alignment = alignment['c_c']

                    ws[f'H{row}'].value = '5'
                    ws[f'H{row}'].font = font['normal']
                    ws[f'H{row}'].alignment = alignment['c_c']

                    ws[f'I{row}'].value = '6'
                    ws[f'I{row}'].font = font['normal']
                    ws[f'I{row}'].alignment = alignment['c_c']


                    ws[f'J{row}'].value = '7'
                    ws[f'J{row}'].font = font['normal']
                    ws[f'J{row}'].alignment = alignment['c_c']

                    make_leftline(ws[f'A{row}'],'thin')
                    make_leftline(ws[f'E{row}'],'thin')
                    make_leftline(ws[f'G{row}'],'thin')
                    make_leftline(ws[f'H{row}'],'thin')
                    make_leftline(ws[f'I{row}'],'thin')
                    make_leftline(ws[f'J{row}'],'thin')
                    make_rightline(ws[f'J{row}'],'thin')

                    for letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'):
                        make_underline(ws[f'{letter}{row}'], 'thin')

                    
                    numero_righe_titoli = 0
                    totalizzatore: int =  0
                    incasso_lordo: float = 0.0 
                    for titolo in titoli:
                        numero_righe_titoli += 1 
                        row = begin_row + 2 + numero_righe_titoli

                        ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row, end_column=colnum('d'))
                        ws[f'A{row}'].value = titolo['tipo']
                        ws[f'A{row}'].font = font['bold']
                        ws[f'A{row}'].alignment = alignment['l_c']

                        try:
                            lettera = titolo['serie_lettera'][0]
                        except:
                            lettera = 'n.d.'
                        ws[f'E{row}'].value = lettera
                        ws[f'E{row}'].font = font['bold']
                        ws[f'E{row}'].alignment = alignment['c_c']

                        try:
                            numero = titolo['serie_numero'][0]
                        except:
                            numero = 'n.d.'
                        ws[f'F{row}'].value = numero
                        ws[f'F{row}'].font = font['bold']
                        ws[f'F{row}'].alignment = alignment['c_c']

                        ws[f'G{row}'].value = titolo['importo']
                        ws[f'G{row}'].font = font['normal']
                        ws[f'G{row}'].alignment = alignment['r_c']
                        ws[f'G{row}'].number_format  = '€ 0.00'

                        ws[f'H{row}'].value = titolo['da']
                        ws[f'H{row}'].font = font['normal']
                        ws[f'H{row}'].alignment = alignment['r_c']

                        ws[f'I{row}'].value = titolo['a']
                        ws[f'I{row}'].font = font['normal']
                        ws[f'I{row}'].alignment = alignment['r_c']


                        ws[f'J{row}'].value = titolo['num']
                        ws[f'J{row}'].font = font['normal']
                        ws[f'J{row}'].alignment = alignment['r_c']

                        totalizzatore += titolo['num']
                        incasso_lordo += titolo['num'] * titolo['importo'] 


                        make_leftline(ws[f'A{row}'],'thin')
                        make_leftline(ws[f'E{row}'],'thin')
                        make_leftline(ws[f'G{row}'],'thin')
                        make_leftline(ws[f'H{row}'],'thin')
                        make_leftline(ws[f'I{row}'],'thin')
                        make_leftline(ws[f'J{row}'],'thin')
                        make_rightline(ws[f'J{row}'],'thin')

                        for letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'):
                            make_underline(ws[f'{letter}{row}'], 'thin')

                    begin_row = row + 1
                    for idx in range(10 - numero_righe_titoli):
                        row = begin_row + idx
                        ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row, end_column=colnum('d'))

                        make_leftline(ws[f'A{row}'],'thin')
                        make_leftline(ws[f'E{row}'],'thin')
                        make_leftline(ws[f'G{row}'],'thin')
                        make_leftline(ws[f'H{row}'],'thin')
                        make_leftline(ws[f'I{row}'],'thin')
                        make_leftline(ws[f'J{row}'],'thin')
                        make_rightline(ws[f'J{row}'],'thin')

                        for letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'):
                            make_underline(ws[f'{letter}{row}'], 'thin')

                    begin_row = row + 1
                    row = begin_row + 0
                    ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row, end_column=colnum('d'))
                    ws[f'A{row}'].value = 'TOTALE BIGLIETTI'
                    ws[f'A{row}'].font = font['normal']
                    ws[f'A{row}'].alignment = alignment['l_c']

                    ws[f'J{row}'].value = totalizzatore
                    ws[f'J{row}'].font = font['bold']
                    ws[f'J{row}'].alignment = alignment['r_c']
                    make_border(ws[f'J{row}'],'thin')
                    
                    row = begin_row + 1
                    ws.row_dimensions[row].height = 5
                
                    row = begin_row + 2
                    ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row, end_column=colnum('d'))
                    ws[f'A{row}'].value = 'INCASSO LORDO'
                    ws[f'A{row}'].font = font['normal']
                    ws[f'A{row}'].alignment = alignment['l_c']

                    ws[f'J{row}'].value = incasso_lordo
                    ws[f'J{row}'].font = font['bold']
                    ws[f'J{row}'].alignment = alignment['r_c']
                    ws[f'J{row}'].number_format  = '€ 0.00'
                    make_border(ws[f'J{row}'],'thin')

                    begin_row = 54

                    row = begin_row + 0

                    for letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'):
                        make_underline(ws[f'{letter}{row}'], 'thin')

                    row = begin_row + 1

                    ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row, end_column=colnum('b'))

                    ws[f'A{row}'].value = 'Contrassegno'
                    ws[f'A{row}'].font = font['normal09']
                    ws[f'A{row}'].alignment = alignment['c_c']
                    make_leftline(ws[f'A{row}'],'thin')
                    make_rightline(ws[f'B{row}'],'thin')

                    ws.merge_cells(start_row=row, start_column=colnum('c'), end_row=row, end_column=colnum('f'))
                    ws[f'C{row}'].value = 'firma per il concessionario'
                    ws[f'C{row}'].font = font['normal09']
                    ws[f'C{row}'].alignment = alignment['c_c']
                    make_leftline(ws[f'C{row}'],'thin')
                    make_rightline(ws[f'F{row}'],'thin')

                    ws.merge_cells(start_row=row, start_column=colnum('g'), end_row=row, end_column=colnum('j'))
                    ws[f'G{row}'].value = 'firma del rappresentante legale'
                    ws[f'G{row}'].font = font['normal09']
                    ws[f'G{row}'].alignment = alignment['c_c']
                    make_leftline(ws[f'G{row}'],'thin')
                    make_rightline(ws[f'J{row}'],'thin')

                    row = begin_row + 2

                    ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row, end_column=colnum('b'))

                    ws[f'A{row}'].value = 'del concessionario(1)'
                    ws[f'A{row}'].font = font['normal09']
                    ws[f'A{row}'].alignment = alignment['c_c']
                    make_leftline(ws[f'A{row}'],'thin')
                    make_rightline(ws[f'B{row}'],'thin')

                    ws.merge_cells(start_row=row, start_column=colnum('c'), end_row=row, end_column=colnum('f'))
                    ws[f'C{row}'].value = ' '
                    ws[f'C{row}'].font = font['normal09']
                    ws[f'C{row}'].alignment = alignment['c_c']
                    make_leftline(ws[f'C{row}'],'thin')
                    make_rightline(ws[f'F{row}'],'thin')

                    ws.merge_cells(start_row=row, start_column=colnum('g'), end_row=row, end_column=colnum('j'))
                    ws[f'G{row}'].value = ''
                    ws[f'G{row}'].font = font['normal09']
                    ws[f'G{row}'].alignment = alignment['c_c']
                    make_leftline(ws[f'G{row}'],'thin')
                    make_rightline(ws[f'J{row}'],'thin')

                    for idx in range(4):
                        row = begin_row +3 + idx
                        ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row, end_column=colnum('b'))
                        make_leftline(ws[f'A{row}'],'thin')
                        make_rightline(ws[f'B{row}'],'thin')
                        ws.merge_cells(start_row=row, start_column=colnum('c'), end_row=row, end_column=colnum('f'))
                        make_leftline(ws[f'C{row}'],'thin')
                        make_rightline(ws[f'F{row}'],'thin')
                        ws.merge_cells(start_row=row, start_column=colnum('g'), end_row=row, end_column=colnum('j'))
                        make_leftline(ws[f'G{row}'],'thin')
                        make_rightline(ws[f'J{row}'],'thin')

                    for letter in ('C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'):
                        make_underline(ws[f'{letter}{row}'], 'thin')

                    row = begin_row +7
                    ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row, end_column=colnum('b'))
                    make_leftline(ws[f'A{row}'],'thin')
                    make_rightline(ws[f'B{row}'],'thin')
                    ws.merge_cells(start_row=row, start_column=colnum('c'), end_row=row, end_column=colnum('f'))
                    make_leftline(ws[f'C{row}'],'thin')
                    make_rightline(ws[f'F{row}'],'thin')
                    ws.merge_cells(start_row=row, start_column=colnum('g'), end_row=row, end_column=colnum('j'))
                    make_leftline(ws[f'G{row}'],'thin')
                    make_rightline(ws[f'J{row}'],'thin')
                    for letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'):
                        make_underline(ws[f'{letter}{row}'], 'thin')
                    
                    row = begin_row +8
                    ws[f'A{row}'].value = "(1) il concessionario è individuato dall'art. 17 del DPR 26 ottobre 1972, n.640 (SIAE)"
                    ws[f'A{row}'].font = font['normal07']
                    ws[f'A{row}'].alignment = alignment['l_c']


                    img_2 = Image(os.path.join(os.getcwd(), save_path,'circle.png'))
                    img_2.height = 200 / 2.2 # insert image height in pixels as float or int (e.g. 305.5)
                    img_2.width= 220 / 2.2 # insert image width in pixels as float or int (e.g. 405.8)
                    img_2.anchor = 'A57' # where you want image to be anchored/start from
                    ws.add_image(img_2) # adding in the image

                except:
                    print("Riquadro Titoli Ingressi - Possibili dati mancanti") 

                wb.save(filename=filename)
                report.updated_at = datetime.now()
                report.save()

            if generate_casher: # Foglio Excel per resoconto cassa

                casher_xlsx_filename = f'{evnt_show_code}.xlsx'
                filename = os.path.join(os.getcwd(), save_path,casher_xlsx_filename)   
                try: 
                    report = Report.objects.get(event_id = curr_event.pk, type ='CASSA')
                    progress_number = report.progress_number
                except:
                    report = Report()
                    report.type = 'CASSA'
                    report.event = curr_event
                    report.doc_path = casher_xlsx_filename
                    report.save()
                    try:
                        all_reports = Report.objects.filter(type = 'CASSA')
                        progress_number_old = 0
                        for report in all_reports:
                            if report.progress_number > progress_number_old:
                                progress_number_old = report.progress_number
                    except:
                        print('No previus CASSA REPORT') 
                    progress_number = progress_number_old + 1
                
                    report.progress_number = progress_number
                    report.save()

                casher_xls_sheetname = f'{evnt_show_code}-{progress_number:03d}'

                try:
                    wb = load_workbook(filename, read_only=False)
                    wss = wb.sheetnames
                    if casher_xls_sheetname in wss:
                        ws = wb[casher_xls_sheetname]
                    else:
                        ws = wb.create_sheet(casher_xls_sheetname)
                        ws.title = casher_xls_sheetname
                except:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = casher_xls_sheetname
                    wb.save(filename)
                
                ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                ws.page_margins.left = 0.25
                ws.page_margins.right = 0.25
                ws.page_margins.top = 0.25
                ws.page_margins.bottom = 0.25
                
                ws.HeaderFooter.differentFirst = True
                ws.firstHeader.left.text = "Page &P of &N"
                ws.firstHeader.center.text = casher_xls_sheetname

                # setting the column size
                colnames = ('A','B','C','D','E','F','G')
                for col in colnames:
                    ws.column_dimensions[col].width = 11

                try: # Dati dello spettacolo INIZIO
                    begin_row = 1
                    row = begin_row + 0

                    ws[f'A{row}'].value = "Data del report:"
                    ws[f'A{row}'].font = font['normal09']
                    ws[f'A{row}'].alignment = alignment['c_c']

                    ws[f'B{row}'].value = fiscal_data.created_at.strftime('%d/%m/%Y')
                    ws[f'B{row}'].font = font['bold']
                    ws[f'B{row}'].alignment = alignment['c_c']

                    ws[f'C{row}'].value = fiscal_data.created_at.strftime('%H:%M')
                    ws[f'C{row}'].font = font['bold']
                    ws[f'C{row}'].alignment = alignment['c_c']

                    make_border(ws[f'B{row}'], 'thin')
                    make_border(ws[f'C{row}'], 'thin')


                    ws[f'D{row}'].value = "Spettacolo:"
                    ws[f'D{row}'].font = font['normal']
                    ws[f'D{row}'].alignment = alignment['r_c']

                    ws.merge_cells(start_row=row, start_column=colnum('e'), end_row=row , end_column=colnum('g'))

                    ws[f'E{row}'].value = curr_event.show.shw_title
                    ws[f'E{row}'].font = font['bold']
                    ws[f'E{row}'].alignment = alignment['c_c']
                    
                    row = begin_row + 1

                    ws[f'A{row}'].value = "Data spettacolo:"
                    ws[f'A{row}'].font = font['normal09']
                    ws[f'A{row}'].alignment = alignment['c_c']

                    date_time_show = curr_event.date_time.astimezone(localtz) 
                    ws[f'B{row}'].value = date_time_show.strftime('%d/%m/%Y')
                    
                    ws[f'B{row}'].font = font['bold']
                    ws[f'B{row}'].alignment = alignment['c_c']

                    ws[f'C{row}'].value = date_time_show.strftime('%H:%M')
                    ws[f'C{row}'].font = font['bold']
                    ws[f'C{row}'].alignment = alignment['c_c']

                    make_border(ws[f'B{row}'], 'thin')
                    make_border(ws[f'C{row}'], 'thin')


                    ws[f'D{row}'].value = "Codice:"
                    ws[f'D{row}'].font = font['normal']
                    ws[f'D{row}'].alignment = alignment['r_c']

                    ws.merge_cells(start_row=row, start_column=colnum('e'), end_row=row , end_column=colnum('g'))

                    ws[f'E{row}'].value = curr_event.show.shw_code
                    ws[f'E{row}'].font = font['bold']
                    ws[f'E{row}'].alignment = alignment['c_c']
                except:
                    print("Errore nella sezione Dati delo spettacolo")
                

                try: # CORRISPETTIVI INIZIO
                    begin_row = 4

                    row = begin_row -1
                    ws.row_dimensions[row].height = 5

                    row = begin_row + 0

                    ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row , end_column=colnum('g'))

                    ws[f'A{row}'].value = "CORRISPETTIVI e INCASSO DA CORRISPETTIVI"
                    ws[f'A{row}'].font = font['bold']
                    ws[f'A{row}'].alignment = alignment['c_c']
                    make_border(ws[f'A{row}'], 'thin')
                    make_rightline(ws[f'G{row}'], 'thin')
                    
                    begin_row = 5
                    row = begin_row + 0
                    ws[f'A{row}'].value = 'TIPO'
                    ws[f'A{row}'].font = font['normal']
                    ws[f'A{row}'].alignment = alignment['c_c']


                    ws[f'B{row}'].value = 'SERIE'
                    ws[f'B{row}'].font = font['normal']
                    ws[f'B{row}'].alignment = alignment['c_c']

                    ws[f'C{row}'].value = 'importo'
                    ws[f'C{row}'].font = font['normal09']
                    ws[f'C{row}'].alignment = alignment['c_c']

                    ws[f'D{row}'].value = 'importo'
                    ws[f'D{row}'].font = font['normal09']
                    ws[f'D{row}'].alignment = alignment['c_c']

                    ws[f'E{row}'].value = 'Dal'
                    ws[f'E{row}'].font = font['normal09']
                    ws[f'E{row}'].alignment = alignment['c_c']

                    ws[f'F{row}'].value = 'Al'
                    ws[f'F{row}'].font = font['normal09']
                    ws[f'F{row}'].alignment = alignment['c_c']


                    ws[f'G{row}'].value = 'Quantità'
                    ws[f'G{row}'].font = font['normal09']
                    ws[f'G{row}'].alignment = alignment['c_c']

                    make_leftline(ws[f'A{row}'],'thin')
                    make_leftline(ws[f'B{row}'],'thin')
                    make_leftline(ws[f'C{row}'],'thin')
                    make_leftline(ws[f'D{row}'],'thin')
                    make_leftline(ws[f'E{row}'],'thin')
                    make_leftline(ws[f'F{row}'],'thin')
                    make_leftline(ws[f'G{row}'],'thin')
                    make_rightline(ws[f'G{row}'],'thin')

                    row = begin_row + 1

                    ws[f'A{row}'].value = ''
                    ws[f'A{row}'].font = font['normal']
                    ws[f'A{row}'].alignment = alignment['c_c']

                    ws[f'B{row}'].value = 'let. & num.'
                    ws[f'B{row}'].font = font['normal09']
                    ws[f'B{row}'].alignment = alignment['c_c']


                    ws[f'C{row}'].value = 'unitario'
                    ws[f'C{row}'].font = font['normal09']
                    ws[f'C{row}'].alignment = alignment['c_c']

                    ws[f'D{row}'].value = 'totale'
                    ws[f'D{row}'].font = font['normal07']
                    ws[f'D{row}'].alignment = alignment['c_c']

                    ws[f'E{row}'].value = 'numero'
                    ws[f'E{row}'].font = font['normal09']
                    ws[f'E{row}'].alignment = alignment['c_c']

                    ws[f'F{row}'].value = 'numero'
                    ws[f'F{row}'].font = font['normal09']
                    ws[f'F{row}'].alignment = alignment['c_c']


                    ws[f'G{row}'].value = '(col.6-col.5+1)'
                    ws[f'G{row}'].font = font['normal07']
                    ws[f'G{row}'].alignment = alignment['c_c']

                    make_leftline(ws[f'A{row}'],'thin')
                    make_leftline(ws[f'B{row}'],'thin')
                    make_leftline(ws[f'C{row}'],'thin')
                    make_leftline(ws[f'D{row}'],'thin')
                    make_leftline(ws[f'E{row}'],'thin')
                    make_leftline(ws[f'F{row}'],'thin')
                    make_leftline(ws[f'G{row}'],'thin')
                    make_rightline(ws[f'G{row}'],'thin')

                    for letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G'):
                        make_underline(ws[f'{letter}{row}'], 'thin')

                    row = begin_row + 2

                    ws[f'A{row}'].value = '1'
                    ws[f'A{row}'].font = font['normal']
                    ws[f'A{row}'].alignment = alignment['c_c']

                    ws[f'B{row}'].value = '2'
                    ws[f'B{row}'].font = font['normal']
                    ws[f'B{row}'].alignment = alignment['c_c']

                    ws[f'C{row}'].value = '3'
                    ws[f'C{row}'].font = font['normal']
                    ws[f'C{row}'].alignment = alignment['c_c']

                    ws[f'D{row}'].value = '4'
                    ws[f'D{row}'].font = font['normal']
                    ws[f'D{row}'].alignment = alignment['c_c']

                    ws[f'E{row}'].value = '5'
                    ws[f'E{row}'].font = font['normal']
                    ws[f'E{row}'].alignment = alignment['c_c']

                    ws[f'F{row}'].value = '6'
                    ws[f'F{row}'].font = font['normal']
                    ws[f'F{row}'].alignment = alignment['c_c']


                    ws[f'G{row}'].value = '7'
                    ws[f'G{row}'].font = font['normal']
                    ws[f'G{row}'].alignment = alignment['c_c']

                    make_leftline(ws[f'A{row}'],'thin')
                    make_leftline(ws[f'B{row}'],'thin')
                    make_leftline(ws[f'C{row}'],'thin')
                    make_leftline(ws[f'D{row}'],'thin')
                    make_leftline(ws[f'E{row}'],'thin')
                    make_leftline(ws[f'F{row}'],'thin')
                    make_leftline(ws[f'G{row}'],'thin')
                    make_rightline(ws[f'G{row}'],'thin')

                    for letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G'):
                        make_underline(ws[f'{letter}{row}'], 'thin')

                    
                    numero_righe_titoli = 0
                    totalizzatore: int =  0
                    incasso_lordo: float = 0.0 
                    for titolo in titoli:
                        numero_righe_titoli += 1 
                        row = begin_row + 2 + numero_righe_titoli

                        ws[f'A{row}'].value = titolo['tipo']
                        ws[f'A{row}'].font = font['bold']
                        ws[f'A{row}'].alignment = alignment['l_c']

                        try:
                            lettera = titolo['serie_lettera'][0]
                        except:
                            lettera = 'n.d.'

                        try:
                            numero = titolo['serie_numero'][0]
                        except:
                            numero = 'n.d.'


                        ws[f'B{row}'].value = f'{lettera}{numero}'
                        ws[f'B{row}'].font = font['bold']
                        ws[f'B{row}'].alignment = alignment['c_c']
 
                        ws[f'C{row}'].value = titolo['importo']
                        ws[f'C{row}'].font = font['bold']
                        ws[f'C{row}'].alignment = alignment['c_c']
                        ws[f'C{row}'].number_format  = '€ 0.00'

                        ws[f'D{row}'].value = titolo['importo'] * titolo['num']
                        ws[f'D{row}'].font = font['normal']
                        ws[f'D{row}'].alignment = alignment['r_c']
                        ws[f'D{row}'].number_format  = '€ 0.00'

                        ws[f'E{row}'].value = titolo['da']
                        ws[f'E{row}'].font = font['normal']
                        ws[f'E{row}'].alignment = alignment['r_c']

                        ws[f'F{row}'].value = titolo['a']
                        ws[f'F{row}'].font = font['normal']
                        ws[f'F{row}'].alignment = alignment['r_c']

                        ws[f'G{row}'].value = titolo['num']
                        ws[f'G{row}'].font = font['normal']
                        ws[f'G{row}'].alignment = alignment['r_c']

                        totalizzatore += titolo['num']
                        incasso_lordo += titolo['num'] * titolo['importo'] 


                        make_leftline(ws[f'A{row}'],'thin')
                        make_leftline(ws[f'B{row}'],'thin')
                        make_leftline(ws[f'C{row}'],'thin')
                        make_leftline(ws[f'D{row}'],'thin')
                        make_leftline(ws[f'E{row}'],'thin')
                        make_leftline(ws[f'F{row}'],'thin')
                        make_leftline(ws[f'G{row}'],'thin')
                        make_rightline(ws[f'G{row}'],'thin')

                        for letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G'):
                            make_underline(ws[f'{letter}{row}'], 'thin')

                    begin_row = row + 1
                    for idx in range(6 - numero_righe_titoli):
                        row = begin_row + idx

                        make_leftline(ws[f'A{row}'],'thin')
                        make_leftline(ws[f'B{row}'],'thin')
                        make_leftline(ws[f'C{row}'],'thin')
                        make_leftline(ws[f'D{row}'],'thin')
                        make_leftline(ws[f'E{row}'],'thin')
                        make_leftline(ws[f'F{row}'],'thin')
                        make_leftline(ws[f'G{row}'],'thin')
                        make_rightline(ws[f'G{row}'],'thin')

                        for letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G'):
                            make_underline(ws[f'{letter}{row}'], 'thin')

                    begin_row = row+1
                    row = begin_row + 0
                    ws[f'A{row}'].value = 'TOTALI'
                    ws[f'A{row}'].font = font['bold']
                    ws[f'A{row}'].alignment = alignment['r_c']

                    ws[f'D{row}'].value = incasso_lordo
                    ws[f'D{row}'].font = font['normal']
                    ws[f'D{row}'].alignment = alignment['r_c']
                    ws[f'D{row}'].number_format  = '€ 0.00'


                    ws[f'G{row}'].value = totalizzatore
                    ws[f'G{row}'].font = font['normal']
                    ws[f'G{row}'].alignment = alignment['r_c']


                    make_leftline(ws[f'A{row}'],'thin')
                    make_rightline(ws[f'A{row}'],'thin')
                    make_leftline(ws[f'D{row}'],'thin')
                    make_rightline(ws[f'D{row}'],'thin')
                    make_leftline(ws[f'G{row}'],'thin')
                    make_rightline(ws[f'G{row}'],'thin')

                    for letter in ('A', 'D',  'G'):
                        make_underline(ws[f'{letter}{row}'], 'thin')

                    row = begin_row + 1
                    ws.row_dimensions[row].height = 5
                except Exception as e:
                    print(f"Errore nella sezione CORRISPETTIVI: {e}")

                try: # CASSA CONTANTI INIZIO
                    statuses = ['begin', 'end']
                    tagli = ['200euro', '100euro', '50euro', '20euro', '10euro', '5euro',
                             '2euro', '1euro', '50cent', '20cent', '10cent']

                    begin_row = row 
                    row = begin_row + 0
                    for letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G'):
                        make_underline(ws[f'{letter}{row}'], 'thin')

                    row = begin_row + 1

                    ws[f'A{row}'].value = "CONTANTI"
                    ws[f'A{row}'].font = font['bold']
                    ws[f'A{row}'].alignment = alignment['c_c']
                    make_border(ws[f'A{row}'], 'thin')

                    ws.merge_cells(start_row=row, start_column=colnum('b'), end_row=row , end_column=colnum('c'))

                    ws[f'B{row}'].value = "CASSA INIZIO"
                    ws[f'B{row}'].font = font['bold']
                    ws[f'B{row}'].alignment = alignment['c_c']
                    make_border(ws[f'B{row}'], 'thin')
                    make_rightline(ws[f'C{row}'], 'thin')

                    ws.merge_cells(start_row=row, start_column=colnum('d'), end_row=row , end_column=colnum('e'))

                    ws[f'D{row}'].value = "CASSA FINE"
                    ws[f'D{row}'].font = font['bold']
                    ws[f'D{row}'].alignment = alignment['c_c']
                    make_border(ws[f'D{row}'], 'thin')
                    make_rightline(ws[f'E{row}'], 'thin')

                    ws.merge_cells(start_row=row, start_column=colnum('f'), end_row=row , end_column=colnum('g'))
                    ws[f'F{row}'].value = "DIFF."
                    ws[f'F{row}'].font = font['bold']
                    ws[f'F{row}'].alignment = alignment['c_c']
                    make_border(ws[f'F{row}'], 'thin')
                    make_rightline(ws[f'G{row}'], 'thin')

                    for letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G'):
                        make_underline(ws[f'{letter}{row}'], 'thin')

                    row = begin_row + 2


                    ws[f'A{row}'].value = "Taglio"
                    ws[f'A{row}'].font = font['normal']
                    ws[f'A{row}'].alignment = alignment['c_c']
                    make_border(ws[f'A{row}'], 'thin')

                    ws[f'B{row}'].value = "Quantità"
                    ws[f'B{row}'].font = font['normal']
                    ws[f'B{row}'].alignment = alignment['c_c']
                    make_border(ws[f'B{row}'], 'thin')

                    ws[f'C{row}'].value = "Importo"
                    ws[f'C{row}'].font = font['normal']
                    ws[f'C{row}'].alignment = alignment['c_c']
                    make_border(ws[f'C{row}'], 'thin')

                    ws[f'D{row}'].value = "Quantità"
                    ws[f'D{row}'].font = font['normal']
                    ws[f'D{row}'].alignment = alignment['c_c']
                    make_border(ws[f'D{row}'], 'thin')

                    ws[f'E{row}'].value = "Importo"
                    ws[f'E{row}'].font = font['normal']
                    ws[f'E{row}'].alignment = alignment['c_c']
                    make_border(ws[f'E{row}'], 'thin')

                    ws[f'F{row}'].value = "Quantità"
                    ws[f'F{row}'].font = font['normal']
                    ws[f'F{row}'].alignment = alignment['c_c']
                    make_border(ws[f'F{row}'], 'thin')

                    ws[f'G{row}'].value = "Euro"
                    ws[f'G{row}'].font = font['normal']
                    ws[f'G{row}'].alignment = alignment['c_c']
                    make_border(ws[f'G{row}'], 'thin')

                    for letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G'):
                        make_underline(ws[f'{letter}{row}'], 'thin')

                    begin_row = row +1
                    numero_tagli_begin = {}
                    importi_tagli_begin = {}
                    diff_tagli = {}
                    totale_begin = 0
                    totale_end = 0
                    differenza_totale = 0
                    for status in statuses:
                        row = begin_row
                        for idx, taglio in enumerate(tagli):
                            row = begin_row + idx 
                            eval_string = f'fiscal_data.cash_{status}_{taglio}'
                            num_pezzi = eval(eval_string)
                            numero_valore_taglio = int(re.findall('[0-9]+',taglio)[0])
                            peso_taglio = re.findall('[a-zA-Z]+',taglio)[0]
                            if peso_taglio == 'euro':
                                valore_taglio = float(numero_valore_taglio)
                            else:
                                valore_taglio = float(numero_valore_taglio / 100)
                            importo_taglio = num_pezzi * valore_taglio
                            if status == 'begin':
                                numero_tagli_begin[taglio] = num_pezzi
                                importi_tagli_begin[taglio] = importo_taglio
                                ws[f'A{row}'].value = f"{numero_valore_taglio} {peso_taglio}"
                                ws[f'A{row}'].font = font['normal']
                                ws[f'A{row}'].alignment = alignment['c_c']
                                make_border(ws[f'A{row}'], 'thin')

                                ws[f'B{row}'].value = num_pezzi
                                ws[f'B{row}'].font = font['normal']
                                ws[f'B{row}'].alignment = alignment['c_c']
                                make_border(ws[f'B{row}'], 'thin')

                                ws[f'C{row}'].value = importo_taglio
                                ws[f'C{row}'].font = font['normal']
                                ws[f'C{row}'].alignment = alignment['r_c']
                                ws[f'C{row}'].number_format  = '€ 0.00'
                                make_border(ws[f'C{row}'], 'thin')

                                totale_begin += importo_taglio

                            else:


                                ws[f'D{row}'].value = num_pezzi
                                ws[f'D{row}'].font = font['normal']
                                ws[f'D{row}'].alignment = alignment['c_c']
                                make_border(ws[f'D{row}'], 'thin')

                                ws[f'E{row}'].value = importo_taglio
                                ws[f'E{row}'].font = font['normal']
                                ws[f'E{row}'].alignment = alignment['r_c']
                                ws[f'E{row}'].number_format  = '€ 0.00'
                                make_border(ws[f'E{row}'], 'thin')

                                diff_numero = num_pezzi - numero_tagli_begin[taglio]
                                diff_euro = importo_taglio - importi_tagli_begin[taglio]
                                differenza_totale += diff_euro
                                diff_tagli[taglio] = diff_euro


                                ws[f'F{row}'].value = diff_numero
                                ws[f'F{row}'].font = font['normal']
                                ws[f'F{row}'].alignment = alignment['c_c']
                                make_border(ws[f'F{row}'], 'thin')

                                ws[f'G{row}'].value = diff_euro
                                ws[f'G{row}'].font = font['normal']
                                ws[f'G{row}'].alignment = alignment['r_c']
                                ws[f'G{row}'].number_format  = '€ 0.00'
                                make_border(ws[f'G{row}'], 'thin')

                                totale_end += importo_taglio

                                del diff_numero, diff_euro

                    begin_row = row + 1
                    row = begin_row + 0

                    ws[f'A{row}'].value = "TOTALI"
                    ws[f'A{row}'].font = font['bold']
                    ws[f'A{row}'].alignment = alignment['c_c']
                    make_border(ws[f'A{row}'], 'thin')


                    ws[f'B{row}'].value = "INIZIO"
                    ws[f'B{row}'].font = font['bold']
                    ws[f'B{row}'].alignment = alignment['c_c']
                    make_border(ws[f'B{row}'], 'thin')

                    ws[f'C{row}'].value = totale_begin
                    ws[f'C{row}'].font = font['normal']
                    ws[f'C{row}'].alignment = alignment['r_c']
                    ws[f'C{row}'].number_format  = '€ 0.00'
                    make_border(ws[f'C{row}'], 'thin')

                    ws[f'D{row}'].value = "FINE"
                    ws[f'D{row}'].font = font['bold']
                    ws[f'D{row}'].alignment = alignment['c_c']
                    make_border(ws[f'D{row}'], 'thin')

                    ws[f'E{row}'].value = totale_end
                    ws[f'E{row}'].font = font['normal']
                    ws[f'E{row}'].alignment = alignment['r_c']
                    ws[f'E{row}'].number_format  = '€ 0.00'
                    make_border(ws[f'E{row}'], 'thin')

                    ws[f'F{row}'].value = "DIFF."
                    ws[f'F{row}'].font = font['bold']
                    ws[f'F{row}'].alignment = alignment['c_c']
                    make_border(ws[f'F{row}'], 'thin')

                    ws[f'G{row}'].value = differenza_totale
                    ws[f'G{row}'].font = font['normal']
                    ws[f'G{row}'].alignment = alignment['r_c']
                    ws[f'G{row}'].number_format  = '€ 0.00'
                    make_border(ws[f'G{row}'], 'thin')

                    row += 1
                    ws.row_dimensions[row].height = 5
                except Exception as e:
                    print(f"Errore nella sezione CASSA CONTANTI: {e}")


                try: # ALTRI INCASSI INIZIO

                    begin_row = row 
                    row = begin_row + 0
                    for letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G'):
                        make_underline(ws[f'{letter}{row}'], 'thin')

                    row = begin_row + 1

                    ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row , end_column=colnum('d'))
                    ws[f'A{row}'].value = "ALTRI INCASSI / METODI di PAGAMENTO"
                    ws[f'A{row}'].font = font['bold']
                    ws[f'A{row}'].alignment = alignment['c_c']
                    make_border(ws[f'A{row}'], 'thin')
                    make_rightline(ws[f'D{row}'], 'thin')

                    row = begin_row + 2

                    ws[f'A{row}'].value = "SATISPAY"
                    ws[f'A{row}'].font = font['bold']
                    ws[f'A{row}'].alignment = alignment['c_c']
                    make_border(ws[f'A{row}'], 'thin')

                    ws[f'B{row}'].value = "SUMUP"
                    ws[f'B{row}'].font = font['bold']
                    ws[f'B{row}'].alignment = alignment['c_c']
                    make_border(ws[f'B{row}'], 'thin')

                    ws[f'C{row}'].value = "CASSA FINE"
                    ws[f'C{row}'].font = font['bold']
                    ws[f'C{row}'].alignment = alignment['c_c']
                    make_border(ws[f'C{row}'], 'thin')

                    ws[f'D{row}'].value = "PREVENDITA"
                    ws[f'D{row}'].font = font['bold']
                    ws[f'D{row}'].alignment = alignment['c_c']
                    make_border(ws[f'D{row}'], 'thin')

                    row = begin_row + 3

                    ws[f'A{row}'].value = fiscal_data.collection_satispay
                    ws[f'A{row}'].font = font['normal']
                    ws[f'A{row}'].alignment = alignment['c_c']
                    ws[f'A{row}'].number_format  = '€ 0.00'
                    make_border(ws[f'A{row}'], 'thin')

                    ws[f'B{row}'].value = fiscal_data.collection_sumup
                    ws[f'B{row}'].font = font['normal']
                    ws[f'B{row}'].alignment = alignment['c_c']
                    ws[f'B{row}'].number_format  = '€ 0.00'
                    make_border(ws[f'B{row}'], 'thin')

                    ws[f'C{row}'].value = fiscal_data.collection_others
                    ws[f'C{row}'].font = font['normal']
                    ws[f'C{row}'].alignment = alignment['r_c']
                    ws[f'C{row}'].number_format  = '€ 0.00'
                    make_border(ws[f'C{row}'], 'thin')

                    ws[f'D{row}'].value = fiscal_data.collection_presell
                    ws[f'D{row}'].font = font['normal']
                    ws[f'D{row}'].alignment = alignment['r_c']
                    ws[f'D{row}'].number_format  = '€ 0.00'
                    make_border(ws[f'D{row}'], 'thin')
                except Exception as e:
                    print(f"Errore nella sezione ALTRI INCASSI: {e}")


                try: # VERIFICHE CONTABILI INIZIO

                    begin_row = row + 2
                    row = begin_row + 0
                    for letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G'):
                        make_underline(ws[f'{letter}{row}'], 'thin')

                    row = begin_row + 1

                    ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row , end_column=colnum('e'))
                    ws[f'A{row}'].value = "VERIFICHE CONTABILI"
                    ws[f'A{row}'].font = font['bold']
                    ws[f'A{row}'].alignment = alignment['c_c']
                    make_border(ws[f'A{row}'], 'thin')
                    make_rightline(ws[f'E{row}'], 'thin')

                    row = begin_row + 2

                    ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row + 3 , end_column=colnum('a'))
                    ws[f'A{row}'].value = "INCASSI"
                    ws[f'A{row}'].font = font['bold']
                    ws[f'A{row}'].alignment = alignment['c_c']
                    make_border(ws[f'A{row}'], 'thin')

                    ws.merge_cells(start_row=row, start_column=colnum('b'), end_row=row  , end_column=colnum('c'))

                    ws[f'B{row}'].value = "CONTANTI"
                    ws[f'B{row}'].font = font['bold']
                    ws[f'B{row}'].alignment = alignment['c_c']
                    make_border(ws[f'B{row}'], 'thin')
                    make_rightline(ws[f'C{row}'], 'thin')

                    ws[f'D{row}'].value = differenza_totale
                    ws[f'D{row}'].font = font['bold']
                    ws[f'D{row}'].alignment = alignment['r_c']
                    ws[f'D{row}'].number_format  = '€ 0.00'
                    make_border(ws[f'D{row}'], 'thin')

                    ws[f'E{row}'].value = "+"
                    ws[f'E{row}'].font = font['bold']
                    ws[f'E{row}'].alignment = alignment['l_c']

                    row = begin_row + 3


                    ws.merge_cells(start_row=row, start_column=colnum('b'), end_row=row  , end_column=colnum('c'))

                    ws[f'B{row}'].value = "ELETTRONICI IN CASSA"
                    ws[f'B{row}'].font = font['bold']
                    ws[f'B{row}'].alignment = alignment['c_c']
                    make_border(ws[f'B{row}'], 'thin')
                    make_rightline(ws[f'C{row}'], 'thin')

                    ws[f'D{row}'].value = fiscal_data.collection_satispay + fiscal_data.collection_sumup
                    ws[f'D{row}'].font = font['bold']
                    ws[f'D{row}'].alignment = alignment['r_c']
                    ws[f'D{row}'].number_format  = '€ 0.00'
                    make_border(ws[f'D{row}'], 'thin')

                    ws[f'E{row}'].value = "+"
                    ws[f'E{row}'].font = font['bold']
                    ws[f'E{row}'].alignment = alignment['l_c']


                    row = begin_row + 4

                    ws.merge_cells(start_row=row, start_column=colnum('b'), end_row=row  , end_column=colnum('c'))

                    ws[f'B{row}'].value = "ALTRI"
                    ws[f'B{row}'].font = font['bold']
                    ws[f'B{row}'].alignment = alignment['c_c']
                    make_border(ws[f'B{row}'], 'thin')
                    make_rightline(ws[f'C{row}'], 'thin')

                    ws[f'D{row}'].value = fiscal_data.collection_others
                    ws[f'D{row}'].font = font['bold']
                    ws[f'D{row}'].alignment = alignment['r_c']
                    ws[f'D{row}'].number_format  = '€ 0.00'
                    make_border(ws[f'D{row}'], 'thin')

                    ws[f'E{row}'].value = "+"
                    ws[f'E{row}'].font = font['bold']
                    ws[f'E{row}'].alignment = alignment['l_c']

                    row = begin_row + 5
                    ws.merge_cells(start_row=row, start_column=colnum('b'), end_row=row  , end_column=colnum('c'))

                    ws[f'B{row}'].value = "Prevendita"
                    ws[f'B{row}'].font = font['bold']
                    ws[f'B{row}'].alignment = alignment['c_c']
                    make_border(ws[f'B{row}'], 'thin')
                    make_rightline(ws[f'C{row}'], 'thin')

                    ws[f'D{row}'].value = fiscal_data.collection_presell
                    ws[f'D{row}'].font = font['bold']
                    ws[f'D{row}'].alignment = alignment['r_c']
                    ws[f'D{row}'].number_format  = '€ 0.00'
                    make_border(ws[f'D{row}'], 'thin')

                    ws[f'E{row}'].value = "+"
                    ws[f'E{row}'].font = font['bold']
                    ws[f'E{row}'].alignment = alignment['l_c']

                    row = begin_row + 6
                    ws.merge_cells(start_row=row, start_column=colnum('b'), end_row=row  , end_column=colnum('c'))

                    ws[f'B{row}'].value = "Incasso Totale"
                    ws[f'B{row}'].font = font['bold']
                    ws[f'B{row}'].alignment = alignment['c_c']
                    make_border(ws[f'B{row}'], 'thin')
                    make_rightline(ws[f'C{row}'], 'thin')

                    incasso_lordo_totale_pagamenti = fiscal_data.collection_presell + fiscal_data.collection_sumup + fiscal_data.collection_others + fiscal_data.collection_satispay + differenza_totale

                    ws[f'D{row}'].value = incasso_lordo_totale_pagamenti
                    ws[f'D{row}'].font = font['bold']
                    ws[f'D{row}'].alignment = alignment['r_c']
                    ws[f'D{row}'].number_format  = '€ 0.00'
                    make_border(ws[f'D{row}'], 'thin')

                    ws[f'E{row}'].value = "="
                    ws[f'E{row}'].font = font['bold']
                    ws[f'E{row}'].alignment = alignment['l_c']


                    row = begin_row + 8

                    ws[f'A{row}'].value = ""
                    ws[f'A{row}'].font = font['bold']
                    ws[f'A{row}'].alignment = alignment['c_c']
                    make_border(ws[f'A{row}'], 'thin')

                    ws.merge_cells(start_row=row, start_column=colnum('b'), end_row=row  , end_column=colnum('c'))

                    ws[f'B{row}'].value = "DA CORRISPETTIVI"
                    ws[f'B{row}'].font = font['bold']
                    ws[f'B{row}'].alignment = alignment['c_c']
                    make_border(ws[f'B{row}'], 'thin')
                    make_rightline(ws[f'C{row}'], 'thin')

                    ws[f'D{row}'].value = incasso_lordo
                    ws[f'D{row}'].font = font['bold']
                    ws[f'D{row}'].alignment = alignment['r_c']
                    ws[f'D{row}'].number_format  = '€ 0.00'
                    make_border(ws[f'D{row}'], 'thin')

                    ws[f'E{row}'].value = "-"
                    ws[f'E{row}'].font = font['bold']
                    ws[f'E{row}'].alignment = alignment['l_c']

                    row = begin_row + 10

                    ws[f'A{row}'].value = "VERIFICA"
                    ws[f'A{row}'].font = font['bold']
                    ws[f'A{row}'].alignment = alignment['c_c']
                    make_border(ws[f'A{row}'], 'thin')

                    ws.merge_cells(start_row=row, start_column=colnum('b'), end_row=row  , end_column=colnum('c'))

                    ws[f'B{row}'].value = "PAGAMENTI - CORRISPETTIVI"
                    ws[f'B{row}'].font = font['bold']
                    ws[f'B{row}'].alignment = alignment['c_c']
                    make_border(ws[f'B{row}'], 'thin')
                    make_rightline(ws[f'C{row}'], 'thin')

                    verifica_incassi = incasso_lordo_totale_pagamenti - incasso_lordo

                    ws[f'D{row}'].value = verifica_incassi
                    ws[f'D{row}'].font = font['bold']
                    ws[f'D{row}'].alignment = alignment['r_c']
                    ws[f'D{row}'].number_format  = '€ 0.00'
                    make_border(ws[f'D{row}'], 'thin')

                    if verifica_incassi == 0.0:
                        ws[f'E{row}'].value = "OK!"
                    else:
                        ws[f'E{row}'].value = "!! KO!"
                    ws[f'E{row}'].font = font['bold']
                    ws[f'E{row}'].alignment = alignment['l_c']
                except Exception as e:
                    print(f"Errore nella sezione VERIFICHE CONTABILI: {e}")


                try: # Cassieri INIZIO

                    begin_row = row + 2
                    row = begin_row + 0

                    ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row , end_column=colnum('g'))
                    ws[f'A{row}'].value = "OPERATORI IN CASSA"
                    ws[f'A{row}'].font = font['bold']
                    ws[f'A{row}'].alignment = alignment['c_c']
                    make_border(ws[f'A{row}'], 'thin')
                    make_rightline(ws[f'G{row}'], 'thin')

                    row = begin_row + 3

                    ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row , end_column=colnum('c'))
                    ws[f'A{row}'].value = fiscal_data.casher_01
                    ws[f'A{row}'].font = font['bold']
                    ws[f'A{row}'].alignment = alignment['c_c']

                    ws.merge_cells(start_row=row, start_column=colnum('e'), end_row=row  , end_column=colnum('g'))

                    ws[f'E{row}'].value = fiscal_data.casher_02
                    ws[f'E{row}'].font = font['bold']
                    ws[f'E{row}'].alignment = alignment['c_c']

                    row = begin_row + 4

                    ws.merge_cells(start_row=row, start_column=colnum('a'), end_row=row +1 , end_column=colnum('c'))
                    make_underline( ws[f'A{row+1}'], 'thin')
                    make_underline( ws[f'B{row+1}'], 'thin')
                    make_underline( ws[f'C{row+1}'], 'thin')

                    ws.merge_cells(start_row=row, start_column=colnum('e'), end_row=row +1  , end_column=colnum('g'))
                    make_underline( ws[f'E{row+1}'], 'thin')
                    make_underline( ws[f'F{row+1}'], 'thin')
                    make_underline( ws[f'G{row+1}'], 'thin')


                except Exception as e:
                    print(f"Errore nella sezione Cassieri: {e}")



                wb.save(filename)
                

            fiscal_data.printed = True
            fiscal_data.save()

            paginator_past = Paginator(past_events, 10)
            page = request.GET.get('page')
            paged_events = paginator_past.get_page(page)



                
            
            context = {
                'past_events' : paged_events,

            }
                    



            # messages.success(request,"Sei entrato con utente di staff! Sei abilitato ad operare.")
            return render(request, 'fiscalmgm/event_list.html', context)
        else:

            
            form = FiscalDataForm(instance=fiscal_data)

            context = {
                'event' : curr_event,
                'form' : form,

            }
            if report_done:
                reports = Report.objects.filter(event_id = curr_event.pk)
                context['reports'] = reports
                messages.warning(request,"I report dell'evento scelto esistono! Potrebbero essere stati già utilizzati.")
            return render(request, 'fiscalmgm/event_reportdata.html', context)

    
@login_required(login_url='login')
def ver_repo(request, event_id):
    if request.user.is_staff:

        curr_event =  Event.objects.get(id = event_id)
        context = {
        'event' : curr_event,            
        }
        try:
            reports = Report.objects.filter(event_id = curr_event.pk)
            context['reports'] = reports
        except:
            messages.warning(request,'Non ci sono report per questo evento')

        return render(request, 'fiscalmgm/event_reportverify.html', context)
    
def make_border(cells_range, style):
    try:
        len(cells_range)

        r = 0
        rows = len(cells_range)
        for row in cells_range:
            r += 1
            if r == 1:
                c=0
                cols = len(row)
                for cell in row:
                    c += 1
                    if c == 1:
                        cell.border =  Border(left=Side(border_style=style, color='FF000000'),right=Side(border_style=None,color='FF000000'),
                                top=Side(border_style=style,color='FF000000'),bottom=Side(border_style=None,color='FF000000'),
                                diagonal=Side(border_style=None,color='FF000000'),diagonal_direction=0,outline=Side(border_style=None,color='FF000000'),
                                vertical=Side(border_style=None,color='FF000000'),horizontal=Side(border_style=None,color='FF000000'))
                    elif c > 1 and c < cols:
                        cell.border =  Border(left=Side(border_style=None, color='FF000000'),right=Side(border_style=None,color='FF000000'),
                                top=Side(border_style=style,color='FF000000'),bottom=Side(border_style=None,color='FF000000'),
                                diagonal=Side(border_style=None,color='FF000000'),diagonal_direction=0,outline=Side(border_style=None,color='FF000000'),
                                vertical=Side(border_style=None,color='FF000000'),horizontal=Side(border_style=None,color='FF000000'))
                    else:
                        cell.border =  Border(left=Side(border_style=None, color='FF000000'),right=Side(border_style=style,color='FF000000'),
                                top=Side(border_style=style,color='FF000000'),bottom=Side(border_style=None,color='FF000000'),
                                diagonal=Side(border_style=None,color='FF000000'),diagonal_direction=0,outline=Side(border_style=None,color='FF000000'),
                                vertical=Side(border_style=None,color='FF000000'),horizontal=Side(border_style=None,color='FF000000'))
            elif r > 1 and r < rows:
                initial_cell = row[0]
                initial_cell.border =  Border(left=Side(border_style=style, color='FF000000'),right=Side(border_style=None,color='FF000000'),
                                top=Side(border_style=None,color='FF000000'),bottom=Side(border_style=None,color='FF000000'),
                                diagonal=Side(border_style=None,color='FF000000'),diagonal_direction=0,outline=Side(border_style=None,color='FF000000'),
                                vertical=Side(border_style=None,color='FF000000'),horizontal=Side(border_style=None,color='FF000000'))
                final_cell = row[-1]
                final_cell.border =  Border(left=Side(border_style=None, color='FF000000'),right=Side(border_style=style,color='FF000000'),
                                top=Side(border_style=None,color='FF000000'),bottom=Side(border_style=None,color='FF000000'),
                                diagonal=Side(border_style=None,color='FF000000'),diagonal_direction=0,outline=Side(border_style=None,color='FF000000'),
                                vertical=Side(border_style=None,color='FF000000'),horizontal=Side(border_style=None,color='FF000000'))
            else:
                c = 0
                for cell in row:
                    c += 1
                    if c == 1:
                        cell.border =  Border(left=Side(border_style=style, color='FF000000'),right=Side(border_style=None,color='FF000000'),
                                top=Side(border_style=None,color='FF000000'),bottom=Side(border_style=style,color='FF000000'),
                                diagonal=Side(border_style=None,color='FF000000'),diagonal_direction=0,outline=Side(border_style=None,color='FF000000'),
                                vertical=Side(border_style=None,color='FF000000'),horizontal=Side(border_style=None,color='FF000000'))
                    elif c > 1 and c < cols:
                        cell.border =  Border(left=Side(border_style=None, color='FF000000'),right=Side(border_style=None,color='FF000000'),
                                top=Side(border_style=None,color='FF000000'),bottom=Side(border_style=style,color='FF000000'),
                                diagonal=Side(border_style=None,color='FF000000'),diagonal_direction=0,outline=Side(border_style=None,color='FF000000'),
                                vertical=Side(border_style=None,color='FF000000'),horizontal=Side(border_style=None,color='FF000000'))
                    else:
                        cell.border =  Border(left=Side(border_style=None, color='FF000000'),right=Side(border_style=style,color='FF000000'),
                                top=Side(border_style=None,color='FF000000'),bottom=Side(border_style=style,color='FF000000'),
                                diagonal=Side(border_style=None,color='FF000000'),diagonal_direction=0,outline=Side(border_style=None,color='FF000000'),
                                vertical=Side(border_style=None,color='FF000000'),horizontal=Side(border_style=None,color='FF000000'))

    except:
        cell = cells_range
        cell.border =  Border(left=Side(border_style=style, color='FF000000'),right=Side(border_style=style,color='FF000000'),
                                top=Side(border_style=style,color='FF000000'),bottom=Side(border_style=style,color='FF000000'),
                                diagonal=Side(border_style=None,color='FF000000'),diagonal_direction=0,outline=Side(border_style=None,color='FF000000'),
                                vertical=Side(border_style=None,color='FF000000'),horizontal=Side(border_style=None,color='FF000000'))

    return


def make_underline(element,style=None):
    try:
        for cell in element[0]:
            styles = get_styles(cell)
            cell.border = Border(left=Side(border_style=styles['left'], color='FF000000'),right=Side(border_style=styles['right'],color='FF000000'),
                            top=Side(border_style=styles['top'],color='FF000000'),bottom=Side(border_style=style,color='FF000000'),
                            diagonal=Side(border_style=styles['diagonal'],color='FF000000'),diagonal_direction=0,outline=Side(border_style=None,color='FF000000'),
                            vertical=Side(border_style=None,color='FF000000'),horizontal=Side(border_style=None,color='FF000000'))

    except:
        styles = get_styles(element)
        element.border =  Border(left=Side(border_style=styles['left'], color='FF000000'),right=Side(border_style=styles['right'],color='FF000000'),
                            top=Side(border_style=styles['top'],color='FF000000'),bottom=Side(border_style=style,color='FF000000'),
                            diagonal=Side(border_style=styles['diagonal'],color='FF000000'),diagonal_direction=0,outline=Side(border_style=None,color='FF000000'),
                            vertical=Side(border_style=None,color='FF000000'),horizontal=Side(border_style=None,color='FF000000'))

    return


def make_leftline(element,style=None):
    try:
        for cell in element[0]:
            styles = get_styles(cell)
            cell.border = Border(left=Side(border_style=style, color='FF000000'),right=Side(border_style=styles['right'],color='FF000000'),
                            top=Side(border_style=styles['top'],color='FF000000'),bottom=Side(border_style=styles['bottom'],color='FF000000'),
                            diagonal=Side(border_style=styles['diagonal'],color='FF000000'),diagonal_direction=0,outline=Side(border_style=None,color='FF000000'),
                            vertical=Side(border_style=None,color='FF000000'),horizontal=Side(border_style=None,color='FF000000'))

    except:
        styles = get_styles(element)
        element.border =  Border(left=Side(border_style=style, color='FF000000'),right=Side(border_style=styles['right'],color='FF000000'),
                            top=Side(border_style=styles['top'],color='FF000000'),bottom=Side(border_style=styles['bottom'],color='FF000000'),
                            diagonal=Side(border_style=styles['diagonal'],color='FF000000'),diagonal_direction=0,outline=Side(border_style=None,color='FF000000'),
                            vertical=Side(border_style=None,color='FF000000'),horizontal=Side(border_style=None,color='FF000000'))

    return


def make_rightline(element,style=None):
    try:
        for cell in element[0]:
            styles = get_styles(cell)
            cell.border = Border(left=Side(border_style=styles['left'], color='FF000000'),right=Side(border_style=style,color='FF000000'),
                            top=Side(border_style=styles['top'],color='FF000000'),bottom=Side(border_style=styles['bottom'],color='FF000000'),
                            diagonal=Side(border_style=styles['diagonal'],color='FF000000'),diagonal_direction=0,outline=Side(border_style=None,color='FF000000'),
                            vertical=Side(border_style=None,color='FF000000'),horizontal=Side(border_style=None,color='FF000000'))

    except:
        styles = get_styles(element)
        element.border =  Border(left=Side(border_style=styles['left'], color='FF000000'),right=Side(border_style=style,color='FF000000'),
                            top=Side(border_style=styles['top'],color='FF000000'),bottom=Side(border_style=styles['bottom'],color='FF000000'),
                            diagonal=Side(border_style=styles['diagonal'],color='FF000000'),diagonal_direction=0,outline=Side(border_style=None,color='FF000000'),
                            vertical=Side(border_style=None,color='FF000000'),horizontal=Side(border_style=None,color='FF000000'))

    return

def get_styles(cell):

    styles={}
    styles['left'] = cell.border.left.style
    styles['right'] = cell.border.right.style
    styles['top'] = cell.border.top.style
    styles['bottom'] = cell.border.bottom.style
    styles['diagonal'] = cell.border.diagonal.style

    return styles

def colnum(letter):
    return ord(letter.lower()) - 96

